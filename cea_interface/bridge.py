"""
Python Bridge for Excel-Microsimulation Integration.

Reads inputs from Excel, runs the microsimulation, and writes results back.
"""

import sys
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
import json

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.simulation import Simulation, SimulationConfig, run_cea, CEAResults
from src.population import PopulationParams, PopulationGenerator
from src.patient import Treatment
from src.treatment import TREATMENT_EFFECTS, TreatmentEffect


class CEABridge:
    """
    Bridge between Excel interface and Python microsimulation.
    """

    # Cell mappings for reading inputs from Excel
    INPUT_CELLS = {
        # Simulation settings (Inputs sheet)
        "n_patients": ("Inputs", "C6"),
        "time_horizon_years": ("Inputs", "C7"),
        "discount_rate": ("Inputs", "C8"),
        "cost_perspective": ("Inputs", "C9"),
        "random_seed": ("Inputs", "C10"),

        # Population parameters
        "age_mean": ("Inputs", "C13"),
        "age_sd": ("Inputs", "C14"),
        "prop_male": ("Inputs", "C15"),
        "sbp_mean": ("Inputs", "C16"),
        "sbp_sd": ("Inputs", "C17"),
        "egfr_mean": ("Inputs", "C18"),
        "egfr_sd": ("Inputs", "C19"),

        # Comorbidities
        "diabetes_prev": ("Inputs", "C22"),
        "smoker_prev": ("Inputs", "C23"),
        "prior_mi_prev": ("Inputs", "C24"),
        "prior_stroke_prev": ("Inputs", "C25"),
        "heart_failure_prev": ("Inputs", "C26"),

        # IXA-001 parameters
        "ixa_sbp_reduction": ("Inputs", "C29"),
        "ixa_sbp_reduction_sd": ("Inputs", "C30"),
        "ixa_monthly_cost": ("Inputs", "C31"),
        "ixa_discontinuation_rate": ("Inputs", "C32"),

        # Spironolactone parameters
        "spiro_sbp_reduction": ("Inputs", "C35"),
        "spiro_sbp_reduction_sd": ("Inputs", "C36"),
        "spiro_monthly_cost": ("Inputs", "C37"),
        "spiro_discontinuation_rate": ("Inputs", "C38"),
    }

    def __init__(self, excel_path: str):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")
        self.excel_path = excel_path
        self.wb = None
        self.inputs = {}

    def read_inputs(self) -> Dict[str, Any]:
        """Read all inputs from the Excel file."""
        self.wb = load_workbook(self.excel_path)

        for param_name, (sheet_name, cell) in self.INPUT_CELLS.items():
            try:
                ws = self.wb[sheet_name]
                value = ws[cell].value
                self.inputs[param_name] = value
            except Exception as e:
                print(f"Warning: Could not read {param_name} from {sheet_name}!{cell}: {e}")
                self.inputs[param_name] = None

        return self.inputs

    def run_simulation(self) -> Dict[str, Any]:
        """Run the microsimulation with the loaded inputs."""
        if not self.inputs:
            self.read_inputs()

        print("\n" + "=" * 60)
        print("RUNNING MICROSIMULATION")
        print("=" * 60)

        # Build population parameters
        pop_params = PopulationParams(
            n_patients=int(self.inputs.get("n_patients", 1000)),
            age_mean=float(self.inputs.get("age_mean", 62.0)),
            age_sd=float(self.inputs.get("age_sd", 10.0)),
            prop_male=float(self.inputs.get("prop_male", 0.55)),
            sbp_mean=float(self.inputs.get("sbp_mean", 155.0)),
            sbp_sd=float(self.inputs.get("sbp_sd", 15.0)),
            egfr_mean=float(self.inputs.get("egfr_mean", 68.0)),
            egfr_sd=float(self.inputs.get("egfr_sd", 20.0)),
            diabetes_prev=float(self.inputs.get("diabetes_prev", 0.35)),
            smoker_prev=float(self.inputs.get("smoker_prev", 0.15)),
            prior_mi_prev=float(self.inputs.get("prior_mi_prev", 0.10)),
            prior_stroke_prev=float(self.inputs.get("prior_stroke_prev", 0.05)),
            heart_failure_prev=float(self.inputs.get("heart_failure_prev", 0.08)),
            seed=int(self.inputs.get("random_seed", 42)),
        )

        # Update treatment effects if modified
        ixa_effect = TREATMENT_EFFECTS[Treatment.IXA_001]
        if self.inputs.get("ixa_sbp_reduction"):
            ixa_effect.sbp_reduction = float(self.inputs["ixa_sbp_reduction"])
        if self.inputs.get("ixa_sbp_reduction_sd"):
            ixa_effect.sbp_reduction_sd = float(self.inputs["ixa_sbp_reduction_sd"])
        if self.inputs.get("ixa_monthly_cost"):
            ixa_effect.monthly_cost = float(self.inputs["ixa_monthly_cost"])
        if self.inputs.get("ixa_discontinuation_rate"):
            ixa_effect.discontinuation_rate = float(self.inputs["ixa_discontinuation_rate"])

        spiro_effect = TREATMENT_EFFECTS[Treatment.SPIRONOLACTONE]
        if self.inputs.get("spiro_sbp_reduction"):
            spiro_effect.sbp_reduction = float(self.inputs["spiro_sbp_reduction"])
        if self.inputs.get("spiro_sbp_reduction_sd"):
            spiro_effect.sbp_reduction_sd = float(self.inputs["spiro_sbp_reduction_sd"])
        if self.inputs.get("spiro_monthly_cost"):
            spiro_effect.monthly_cost = float(self.inputs["spiro_monthly_cost"])
        if self.inputs.get("spiro_discontinuation_rate"):
            spiro_effect.discontinuation_rate = float(self.inputs["spiro_discontinuation_rate"])

        # Run CEA
        print(f"\nPatients per arm: {pop_params.n_patients}")
        print(f"Time horizon: {self.inputs.get('time_horizon_years', 40)} years")
        print(f"IXA-001 SBP reduction: {ixa_effect.sbp_reduction} mmHg")
        print(f"Spironolactone SBP reduction: {spiro_effect.sbp_reduction} mmHg")
        print("\nRunning simulation (this may take 5-10 minutes)...")

        cea_results = run_cea(
            n_patients=pop_params.n_patients,
            time_horizon_years=int(self.inputs.get("time_horizon_years", 40)),
            seed=pop_params.seed,
            perspective=self.inputs.get("cost_perspective", "US"),
        )

        # Extract results
        results = self._extract_results(cea_results)
        results["timestamp"] = datetime.now().isoformat()

        print("\nSimulation complete!")
        return results

    def _extract_results(self, cea: CEAResults) -> Dict[str, Any]:
        """Extract results from CEA into dictionary."""
        return {
            "icer": cea.icer,
            "incremental_costs": cea.incremental_costs,
            "incremental_qalys": cea.incremental_qalys,

            "ixa_mean_costs": cea.intervention.mean_costs,
            "ixa_mean_qalys": cea.intervention.mean_qalys,
            "ixa_mean_life_years": cea.intervention.mean_life_years,

            "spiro_mean_costs": cea.comparator.mean_costs,
            "spiro_mean_qalys": cea.comparator.mean_qalys,
            "spiro_mean_life_years": cea.comparator.mean_life_years,

            "ixa_mi_events": cea.intervention.mi_events,
            "ixa_stroke_events": cea.intervention.stroke_events,
            "ixa_hf_events": cea.intervention.hf_events,
            "ixa_esrd_events": cea.intervention.esrd_events,
            "ixa_cv_deaths": cea.intervention.cv_deaths,

            "spiro_mi_events": cea.comparator.mi_events,
            "spiro_stroke_events": cea.comparator.stroke_events,
            "spiro_hf_events": cea.comparator.hf_events,
            "spiro_esrd_events": cea.comparator.esrd_events,
            "spiro_cv_deaths": cea.comparator.cv_deaths,

            "strokes_avoided": cea.comparator.stroke_events - cea.intervention.stroke_events,
        }

    def write_results(self, results: Dict[str, Any]):
        """Write results back to the Excel file."""
        if self.wb is None:
            self.wb = load_workbook(self.excel_path)

        ws = self.wb["Results"]

        # Key metrics
        ws["C7"] = results.get("icer")
        ws["C9"] = results.get("ixa_mean_costs")
        ws["C10"] = results.get("spiro_mean_costs")
        ws["C11"] = results.get("incremental_costs")
        ws["C13"] = results.get("ixa_mean_qalys")
        ws["C14"] = results.get("spiro_mean_qalys")
        ws["C15"] = results.get("incremental_qalys")

        # Event counts
        ws["C19"] = results.get("ixa_mi_events")
        ws["D19"] = results.get("spiro_mi_events")
        ws["C20"] = results.get("ixa_stroke_events")
        ws["D20"] = results.get("spiro_stroke_events")
        ws["C21"] = results.get("ixa_hf_events")
        ws["D21"] = results.get("spiro_hf_events")
        ws["C22"] = results.get("ixa_esrd_events")
        ws["D22"] = results.get("spiro_esrd_events")
        ws["C23"] = results.get("ixa_cv_deaths")
        ws["D23"] = results.get("spiro_cv_deaths")

        # Update timestamp
        ws["B3"] = f"Last Run: {results.get('timestamp', datetime.now().isoformat())}"

        # Update interpretation
        icer = results.get("icer", 0)
        if icer and icer < 50000:
            interpretation = "IXA-001 is HIGHLY COST-EFFECTIVE (ICER < $50,000/QALY)"
        elif icer and icer < 100000:
            interpretation = "IXA-001 is COST-EFFECTIVE (ICER < $100,000/QALY)"
        elif icer and icer < 150000:
            interpretation = "IXA-001 is MARGINALLY COST-EFFECTIVE ($100K-$150K/QALY)"
        else:
            interpretation = "IXA-001 may NOT be cost-effective (ICER > $150,000/QALY)"
        ws["B26"] = interpretation

        self.wb.save(self.excel_path)
        print(f"\nResults written to: {self.excel_path}")

    def run_and_update(self) -> Dict[str, Any]:
        """Read inputs, run simulation, and write results."""
        self.read_inputs()
        results = self.run_simulation()
        self.write_results(results)
        return results


def run_from_excel(excel_path: str) -> Dict[str, Any]:
    """
    Convenience function to run CEA from Excel file.

    Args:
        excel_path: Path to Excel file with inputs

    Returns:
        Dictionary of results
    """
    bridge = CEABridge(excel_path)
    return bridge.run_and_update()


def print_results(results: Dict[str, Any]):
    """Print formatted results."""
    print("\n" + "=" * 60)
    print("COST-EFFECTIVENESS ANALYSIS RESULTS")
    print("=" * 60)

    print(f"\n{'Metric':<35} {'IXA-001':>15} {'Spironolactone':>15}")
    print("-" * 65)
    print(f"{'Mean Costs':<35} ${results['ixa_mean_costs']:>14,.0f} ${results['spiro_mean_costs']:>14,.0f}")
    print(f"{'Mean QALYs':<35} {results['ixa_mean_qalys']:>15.3f} {results['spiro_mean_qalys']:>15.3f}")

    print("-" * 65)
    print("EVENT COUNTS (per 1,000 patients)")
    print(f"{'Myocardial Infarction':<35} {results['ixa_mi_events']:>15} {results['spiro_mi_events']:>15}")
    print(f"{'Stroke':<35} {results['ixa_stroke_events']:>15} {results['spiro_stroke_events']:>15}")
    print(f"{'Heart Failure':<35} {results['ixa_hf_events']:>15} {results['spiro_hf_events']:>15}")
    print(f"{'ESRD':<35} {results['ixa_esrd_events']:>15} {results['spiro_esrd_events']:>15}")
    print(f"{'CV Death':<35} {results['ixa_cv_deaths']:>15} {results['spiro_cv_deaths']:>15}")

    print("-" * 65)
    print(f"{'Incremental Costs:':<35} ${results['incremental_costs']:>14,.0f}")
    print(f"{'Incremental QALYs:':<35} {results['incremental_qalys']:>15.3f}")

    if results['icer']:
        print(f"{'ICER ($/QALY):':<35} ${results['icer']:>14,.0f}")

        if results['icer'] < 50000:
            interpretation = "HIGHLY COST-EFFECTIVE (< $50K/QALY)"
        elif results['icer'] < 100000:
            interpretation = "COST-EFFECTIVE (< $100K/QALY)"
        elif results['icer'] < 150000:
            interpretation = "MARGINALLY COST-EFFECTIVE ($100-150K/QALY)"
        else:
            interpretation = "NOT COST-EFFECTIVE (> $150K/QALY)"

        print(f"\nInterpretation: {interpretation}")

    print("=" * 60)
