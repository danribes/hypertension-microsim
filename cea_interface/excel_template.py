"""
Excel Template Generator for CEA Microsimulation.

Creates a user-friendly Excel workbook for specifying model inputs
and viewing results from the microsimulation.
"""

from typing import Dict, Any, Optional
from pathlib import Path
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle, Protection
    )
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class CEAExcelTemplate:
    """
    Generates Excel template for CEA model inputs and results.
    """

    # Style constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    CALC_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Default parameters
    DEFAULT_PARAMS = {
        # Simulation settings
        "n_patients": 1000,
        "time_horizon_years": 40,
        "discount_rate": 0.03,
        "cost_perspective": "US",
        "random_seed": 42,

        # Population parameters
        "age_mean": 62.0,
        "age_sd": 10.0,
        "prop_male": 0.55,
        "sbp_mean": 155.0,
        "sbp_sd": 15.0,
        "egfr_mean": 68.0,
        "egfr_sd": 20.0,

        # Comorbidity prevalence
        "diabetes_prev": 0.35,
        "smoker_prev": 0.15,
        "prior_mi_prev": 0.10,
        "prior_stroke_prev": 0.05,
        "heart_failure_prev": 0.08,

        # IXA-001 parameters
        "ixa_sbp_reduction": 20.0,
        "ixa_sbp_reduction_sd": 8.0,
        "ixa_monthly_cost": 500.0,
        "ixa_discontinuation_rate": 0.12,

        # Spironolactone parameters
        "spiro_sbp_reduction": 9.0,
        "spiro_sbp_reduction_sd": 6.0,
        "spiro_monthly_cost": 15.0,
        "spiro_discontinuation_rate": 0.15,

        # Utility values
        "utility_controlled_htn": 0.85,
        "utility_uncontrolled_htn": 0.80,
        "disutility_mi": 0.05,
        "disutility_stroke": 0.15,
        "disutility_hf": 0.10,
        "disutility_esrd": 0.20,

        # Event costs (US)
        "cost_mi_acute": 25000.0,
        "cost_stroke_acute": 15200.0,
        "cost_hf_admission": 18000.0,
        "cost_esrd_annual": 90000.0,
    }

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Set up named styles."""
        header_style = NamedStyle(name="header_style")
        header_style.font = self.HEADER_FONT
        header_style.fill = self.HEADER_FILL
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = self.BORDER
        self.wb.add_named_style(header_style)

        input_style = NamedStyle(name="input_style")
        input_style.fill = self.INPUT_FILL
        input_style.border = self.BORDER
        input_style.alignment = Alignment(horizontal="right")
        input_style.protection = Protection(locked=False)
        self.wb.add_named_style(input_style)

        result_style = NamedStyle(name="result_style")
        result_style.fill = self.RESULT_FILL
        result_style.border = self.BORDER
        result_style.alignment = Alignment(horizontal="right")
        result_style.font = Font(bold=True)
        self.wb.add_named_style(result_style)

    def generate(self, output_path: str, precomputed_results: Optional[Dict] = None) -> str:
        """Generate the Excel template."""
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        self._create_instructions_sheet()
        self._create_inputs_sheet()
        self._create_results_sheet(precomputed_results)
        self._create_scenarios_sheet(precomputed_results)
        self._create_sensitivity_sheet(precomputed_results)

        self.wb.active = self.wb["Instructions"]
        self.wb.save(output_path)
        return output_path

    def _create_instructions_sheet(self):
        """Create instructions sheet."""
        ws = self.wb.create_sheet("Instructions", 0)

        ws.merge_cells("B2:H2")
        ws["B2"] = "IXA-001 COST-EFFECTIVENESS MODEL"
        ws["B2"].font = Font(size=20, bold=True, color="1F4E79")

        ws["B3"] = "Hybrid Excel-Python Interface"
        ws["B3"].font = Font(size=14, italic=True, color="666666")

        instructions = [
            "",
            "HOW TO USE THIS MODEL",
            "=" * 50,
            "",
            "OPTION 1: USE PRE-COMPUTED SCENARIOS (Instant Results)",
            "-" * 50,
            "1. Go to 'Scenarios' sheet to see pre-computed results",
            "2. Results available for: Base Case, Price Sensitivity, Subgroups",
            "3. No waiting - results are already calculated!",
            "",
            "OPTION 2: RUN CUSTOM SCENARIOS (5-10 minutes)",
            "-" * 50,
            "1. Go to 'Inputs' sheet and modify YELLOW cells",
            "2. Save this Excel file",
            "3. Run the Python bridge script:",
            "   python run_cea_from_excel.py --input THIS_FILE.xlsx",
            "4. Results will be written back to this file",
            "5. Open the updated file to see your results",
            "",
            "WHAT YOU CAN MODIFY",
            "-" * 50,
            "- Population: Age, sex distribution, comorbidities",
            "- Treatment: SBP reduction, costs, discontinuation rates",
            "- Costs: Event costs, annual management costs",
            "- Utilities: Health state preferences (QALYs)",
            "- Settings: Time horizon, discount rate, sample size",
            "",
            "MODEL OUTPUTS",
            "-" * 50,
            "- ICER (Incremental Cost-Effectiveness Ratio)",
            "- Total Costs and QALYs per patient",
            "- Event counts (MI, Stroke, HF, ESRD)",
            "- Sensitivity analysis results",
            "",
            "COLOR LEGEND",
            "-" * 50,
            "YELLOW = Editable input (modify these)",
            "GREEN = Results (automatically calculated)",
            "BLUE = Pre-computed scenarios",
            "GRAY = Locked/system values",
            "",
            "IMPORTANT NOTES",
            "-" * 50,
            "- Each simulation run takes 5-10 minutes",
            "- Use pre-computed scenarios when possible for speed",
            "- Results are based on 1,000 patient Monte Carlo simulation",
            "- ICER < $100,000/QALY is considered cost-effective (US)",
        ]

        for i, line in enumerate(instructions):
            ws[f"B{5+i}"] = line
            if line and not line.startswith("-") and not line.startswith("="):
                if line == line.upper() or line.startswith("OPTION"):
                    ws[f"B{5+i}"].font = Font(bold=True, size=11, color="1F4E79")

        ws.column_dimensions["B"].width = 70

    def _create_inputs_sheet(self):
        """Create the inputs sheet with all modifiable parameters."""
        ws = self.wb.create_sheet("Inputs")

        ws.merge_cells("B2:E2")
        ws["B2"] = "MODEL INPUTS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        ws["B3"] = "Modify YELLOW cells, save file, then run Python script"
        ws["B3"].font = Font(size=11, italic=True, color="E67300")

        # ========== SIMULATION SETTINGS ==========
        row = 5
        ws[f"B{row}"] = "SIMULATION SETTINGS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        settings = [
            ("Number of Patients", "n_patients", self.DEFAULT_PARAMS["n_patients"], "patients", "#,##0"),
            ("Time Horizon", "time_horizon_years", self.DEFAULT_PARAMS["time_horizon_years"], "years", "0"),
            ("Discount Rate", "discount_rate", self.DEFAULT_PARAMS["discount_rate"], "", "0.0%"),
            ("Cost Perspective", "cost_perspective", self.DEFAULT_PARAMS["cost_perspective"], "(US/UK)", "@"),
            ("Random Seed", "random_seed", self.DEFAULT_PARAMS["random_seed"], "", "0"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(settings):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== POPULATION PARAMETERS ==========
        row = row + len(settings) + 2
        ws[f"B{row}"] = "POPULATION PARAMETERS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        population = [
            ("Age - Mean", "age_mean", self.DEFAULT_PARAMS["age_mean"], "years", "0.0"),
            ("Age - SD", "age_sd", self.DEFAULT_PARAMS["age_sd"], "years", "0.0"),
            ("Proportion Male", "prop_male", self.DEFAULT_PARAMS["prop_male"], "", "0.0%"),
            ("Baseline SBP - Mean", "sbp_mean", self.DEFAULT_PARAMS["sbp_mean"], "mmHg", "0.0"),
            ("Baseline SBP - SD", "sbp_sd", self.DEFAULT_PARAMS["sbp_sd"], "mmHg", "0.0"),
            ("Baseline eGFR - Mean", "egfr_mean", self.DEFAULT_PARAMS["egfr_mean"], "mL/min", "0.0"),
            ("Baseline eGFR - SD", "egfr_sd", self.DEFAULT_PARAMS["egfr_sd"], "mL/min", "0.0"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(population):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== COMORBIDITIES ==========
        row = row + len(population) + 2
        ws[f"B{row}"] = "COMORBIDITY PREVALENCE"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        comorbidities = [
            ("Diabetes", "diabetes_prev", self.DEFAULT_PARAMS["diabetes_prev"], "", "0.0%"),
            ("Current Smoker", "smoker_prev", self.DEFAULT_PARAMS["smoker_prev"], "", "0.0%"),
            ("Prior MI", "prior_mi_prev", self.DEFAULT_PARAMS["prior_mi_prev"], "", "0.0%"),
            ("Prior Stroke", "prior_stroke_prev", self.DEFAULT_PARAMS["prior_stroke_prev"], "", "0.0%"),
            ("Heart Failure", "heart_failure_prev", self.DEFAULT_PARAMS["heart_failure_prev"], "", "0.0%"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(comorbidities):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== IXA-001 PARAMETERS ==========
        row = row + len(comorbidities) + 2
        ws[f"B{row}"] = "IXA-001 TREATMENT PARAMETERS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        ixa_params = [
            ("SBP Reduction - Mean", "ixa_sbp_reduction", self.DEFAULT_PARAMS["ixa_sbp_reduction"], "mmHg", "0.0"),
            ("SBP Reduction - SD", "ixa_sbp_reduction_sd", self.DEFAULT_PARAMS["ixa_sbp_reduction_sd"], "mmHg", "0.0"),
            ("Monthly Drug Cost", "ixa_monthly_cost", self.DEFAULT_PARAMS["ixa_monthly_cost"], "$", '"$"#,##0'),
            ("Annual Discontinuation", "ixa_discontinuation_rate", self.DEFAULT_PARAMS["ixa_discontinuation_rate"], "", "0.0%"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(ixa_params):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== SPIRONOLACTONE PARAMETERS ==========
        row = row + len(ixa_params) + 2
        ws[f"B{row}"] = "SPIRONOLACTONE TREATMENT PARAMETERS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        spiro_params = [
            ("SBP Reduction - Mean", "spiro_sbp_reduction", self.DEFAULT_PARAMS["spiro_sbp_reduction"], "mmHg", "0.0"),
            ("SBP Reduction - SD", "spiro_sbp_reduction_sd", self.DEFAULT_PARAMS["spiro_sbp_reduction_sd"], "mmHg", "0.0"),
            ("Monthly Drug Cost", "spiro_monthly_cost", self.DEFAULT_PARAMS["spiro_monthly_cost"], "$", '"$"#,##0'),
            ("Annual Discontinuation", "spiro_discontinuation_rate", self.DEFAULT_PARAMS["spiro_discontinuation_rate"], "", "0.0%"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(spiro_params):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== UTILITY VALUES ==========
        row = row + len(spiro_params) + 2
        ws[f"B{row}"] = "UTILITY VALUES (QALY Weights)"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        utilities = [
            ("Controlled HTN", "utility_controlled_htn", self.DEFAULT_PARAMS["utility_controlled_htn"], "", "0.00"),
            ("Uncontrolled HTN", "utility_uncontrolled_htn", self.DEFAULT_PARAMS["utility_uncontrolled_htn"], "", "0.00"),
            ("Disutility - MI", "disutility_mi", self.DEFAULT_PARAMS["disutility_mi"], "", "0.00"),
            ("Disutility - Stroke", "disutility_stroke", self.DEFAULT_PARAMS["disutility_stroke"], "", "0.00"),
            ("Disutility - HF", "disutility_hf", self.DEFAULT_PARAMS["disutility_hf"], "", "0.00"),
            ("Disutility - ESRD", "disutility_esrd", self.DEFAULT_PARAMS["disutility_esrd"], "", "0.00"),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(utilities):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # ========== EVENT COSTS ==========
        row = row + len(utilities) + 2
        ws[f"B{row}"] = "EVENT COSTS (US Dollars)"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:E{row}")

        costs = [
            ("MI - Acute Event", "cost_mi_acute", self.DEFAULT_PARAMS["cost_mi_acute"], "", '"$"#,##0'),
            ("Stroke - Acute Event", "cost_stroke_acute", self.DEFAULT_PARAMS["cost_stroke_acute"], "", '"$"#,##0'),
            ("HF Admission", "cost_hf_admission", self.DEFAULT_PARAMS["cost_hf_admission"], "", '"$"#,##0'),
            ("ESRD - Annual", "cost_esrd_annual", self.DEFAULT_PARAMS["cost_esrd_annual"], "", '"$"#,##0'),
        ]

        for i, (label, key, value, unit, fmt) in enumerate(costs):
            r = row + 1 + i
            ws[f"B{r}"] = label
            ws[f"C{r}"] = value
            ws[f"C{r}"].style = "input_style"
            ws[f"C{r}"].number_format = fmt
            ws[f"D{r}"] = unit

        # Set column widths
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10

    def _create_results_sheet(self, precomputed: Optional[Dict] = None):
        """Create results sheet."""
        ws = self.wb.create_sheet("Results")

        ws.merge_cells("B2:F2")
        ws["B2"] = "COST-EFFECTIVENESS RESULTS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        if precomputed and "base_case" in precomputed:
            results = precomputed["base_case"]
            ws["B3"] = f"Last Run: {results.get('timestamp', 'Pre-computed')}"
        else:
            ws["B3"] = "Run Python script to populate results"
        ws["B3"].font = Font(size=11, italic=True, color="666666")

        # ========== KEY METRICS ==========
        row = 5
        ws[f"B{row}"] = "KEY METRICS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        if precomputed and "base_case" in precomputed:
            r = precomputed["base_case"]
            metrics = [
                ("ICER ($/QALY)", r.get("icer", "N/A"), '"$"#,##0'),
                ("", "", ""),
                ("IXA-001 Mean Costs", r.get("ixa_mean_costs", "N/A"), '"$"#,##0'),
                ("Spironolactone Mean Costs", r.get("spiro_mean_costs", "N/A"), '"$"#,##0'),
                ("Incremental Costs", r.get("incremental_costs", "N/A"), '"$"#,##0'),
                ("", "", ""),
                ("IXA-001 Mean QALYs", r.get("ixa_mean_qalys", "N/A"), "0.000"),
                ("Spironolactone Mean QALYs", r.get("spiro_mean_qalys", "N/A"), "0.000"),
                ("Incremental QALYs", r.get("incremental_qalys", "N/A"), "0.000"),
            ]
        else:
            metrics = [
                ("ICER ($/QALY)", "[Run simulation]", "@"),
                ("", "", ""),
                ("IXA-001 Mean Costs", "[Run simulation]", "@"),
                ("Spironolactone Mean Costs", "[Run simulation]", "@"),
                ("Incremental Costs", "[Run simulation]", "@"),
                ("", "", ""),
                ("IXA-001 Mean QALYs", "[Run simulation]", "@"),
                ("Spironolactone Mean QALYs", "[Run simulation]", "@"),
                ("Incremental QALYs", "[Run simulation]", "@"),
            ]

        for i, (label, value, fmt) in enumerate(metrics):
            r = row + 1 + i
            if label:
                ws[f"B{r}"] = label
                ws[f"C{r}"] = value
                if value != "[Run simulation]" and value != "N/A" and value != "":
                    ws[f"C{r}"].style = "result_style"
                ws[f"C{r}"].number_format = fmt

        # ========== EVENT COUNTS ==========
        row = row + len(metrics) + 2
        ws[f"B{row}"] = "EVENT COUNTS (per 1,000 patients)"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        headers = ["Event", "IXA-001", "Spironolactone"]
        for i, header in enumerate(headers):
            ws.cell(row=row+1, column=2+i, value=header).style = "header_style"

        if precomputed and "base_case" in precomputed:
            r = precomputed["base_case"]
            events = [
                ("Myocardial Infarction", r.get("ixa_mi_events", "N/A"), r.get("spiro_mi_events", "N/A")),
                ("Stroke (Total)", r.get("ixa_stroke_events", "N/A"), r.get("spiro_stroke_events", "N/A")),
                ("Heart Failure", r.get("ixa_hf_events", "N/A"), r.get("spiro_hf_events", "N/A")),
                ("ESRD", r.get("ixa_esrd_events", "N/A"), r.get("spiro_esrd_events", "N/A")),
                ("CV Death", r.get("ixa_cv_deaths", "N/A"), r.get("spiro_cv_deaths", "N/A")),
            ]
        else:
            events = [
                ("Myocardial Infarction", "-", "-"),
                ("Stroke (Total)", "-", "-"),
                ("Heart Failure", "-", "-"),
                ("ESRD", "-", "-"),
                ("CV Death", "-", "-"),
            ]

        for i, (event, ixa, spiro) in enumerate(events):
            r = row + 2 + i
            ws.cell(row=r, column=2, value=event)
            ws.cell(row=r, column=3, value=ixa)
            ws.cell(row=r, column=4, value=spiro)

        # ========== INTERPRETATION ==========
        row = row + len(events) + 4
        ws[f"B{row}"] = "INTERPRETATION"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        if precomputed and "base_case" in precomputed:
            icer = precomputed["base_case"].get("icer", 0)
            if icer and icer < 50000:
                interpretation = "IXA-001 is HIGHLY COST-EFFECTIVE (ICER < $50,000/QALY)"
                color = "228B22"  # Green
            elif icer and icer < 100000:
                interpretation = "IXA-001 is COST-EFFECTIVE (ICER < $100,000/QALY)"
                color = "228B22"  # Green
            elif icer and icer < 150000:
                interpretation = "IXA-001 is MARGINALLY COST-EFFECTIVE ($100K-$150K/QALY)"
                color = "FFA500"  # Orange
            else:
                interpretation = "IXA-001 may NOT be cost-effective (ICER > $150,000/QALY)"
                color = "FF0000"  # Red
        else:
            interpretation = "Run simulation to see interpretation"
            color = "666666"

        ws[f"B{row+1}"] = interpretation
        ws[f"B{row+1}"].font = Font(bold=True, size=12, color=color)

        # Set column widths
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    def _create_scenarios_sheet(self, precomputed: Optional[Dict] = None):
        """Create pre-computed scenarios sheet."""
        ws = self.wb.create_sheet("Scenarios")

        ws.merge_cells("B2:H2")
        ws["B2"] = "PRE-COMPUTED SCENARIOS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        ws["B3"] = "Instant results - no waiting required!"
        ws["B3"].font = Font(size=11, italic=True, color="228B22")

        # ========== SCENARIO TABLE ==========
        row = 5
        ws[f"B{row}"] = "SCENARIO COMPARISON"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:H{row}")

        headers = ["Scenario", "IXA Cost", "ICER", "Incr. Costs", "Incr. QALYs", "Strokes Avoided", "Interpretation"]
        for i, header in enumerate(headers):
            ws.cell(row=row+1, column=2+i, value=header).style = "header_style"

        if precomputed:
            scenarios = []
            for name, data in precomputed.items():
                if isinstance(data, dict) and "icer" in data:
                    icer = data.get("icer", 0)
                    if icer and icer < 100000:
                        interp = "Cost-Effective"
                    elif icer and icer < 150000:
                        interp = "Marginal"
                    else:
                        interp = "Not CE"

                    scenarios.append((
                        data.get("scenario_name", name),
                        data.get("ixa_monthly_cost", 500) * 12,
                        data.get("icer", "N/A"),
                        data.get("incremental_costs", "N/A"),
                        data.get("incremental_qalys", "N/A"),
                        data.get("strokes_avoided", "N/A"),
                        interp
                    ))

            for i, (name, cost, icer, inc_cost, inc_qaly, strokes, interp) in enumerate(scenarios):
                r = row + 2 + i
                ws.cell(row=r, column=2, value=name)
                ws.cell(row=r, column=3, value=cost).number_format = '"$"#,##0'
                ws.cell(row=r, column=4, value=icer).number_format = '"$"#,##0'
                ws.cell(row=r, column=5, value=inc_cost).number_format = '"$"#,##0'
                ws.cell(row=r, column=6, value=inc_qaly).number_format = "0.000"
                ws.cell(row=r, column=7, value=strokes)
                ws.cell(row=r, column=8, value=interp)

                if interp == "Cost-Effective":
                    ws.cell(row=r, column=8).font = Font(color="228B22", bold=True)
                elif interp == "Marginal":
                    ws.cell(row=r, column=8).font = Font(color="FFA500", bold=True)
                else:
                    ws.cell(row=r, column=8).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(row=row+2, column=2, value="[Pre-computed scenarios will appear here]")

        # Set column widths
        ws.column_dimensions["B"].width = 25
        for col in range(3, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_sensitivity_sheet(self, precomputed: Optional[Dict] = None):
        """Create sensitivity analysis sheet."""
        ws = self.wb.create_sheet("Sensitivity")

        ws.merge_cells("B2:F2")
        ws["B2"] = "SENSITIVITY ANALYSIS"
        ws["B2"].font = Font(size=16, bold=True, color="1F4E79")

        # ========== PRICE SENSITIVITY ==========
        row = 4
        ws[f"B{row}"] = "IXA-001 PRICE SENSITIVITY"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        headers = ["Monthly Price", "Annual Cost", "ICER"]
        for i, header in enumerate(headers):
            ws.cell(row=row+1, column=2+i, value=header).style = "header_style"

        if precomputed and "price_sensitivity" in precomputed:
            for i, result in enumerate(precomputed["price_sensitivity"]):
                r = row + 2 + i
                ws.cell(row=r, column=2, value=result["monthly_price"]).number_format = '"$"#,##0'
                ws.cell(row=r, column=3, value=result["annual_cost"]).number_format = '"$"#,##0'
                ws.cell(row=r, column=4, value=result["icer"]).number_format = '"$"#,##0'
        else:
            prices = [300, 400, 500, 600, 700, 800]
            for i, price in enumerate(prices):
                r = row + 2 + i
                ws.cell(row=r, column=2, value=price).number_format = '"$"#,##0'
                ws.cell(row=r, column=3, value=price * 12).number_format = '"$"#,##0'
                ws.cell(row=r, column=4, value="[Run scenarios]")

        # ========== SUBGROUP ANALYSIS ==========
        row = row + 10
        ws[f"B{row}"] = "SUBGROUP ANALYSIS"
        ws[f"B{row}"].style = "header_style"
        ws.merge_cells(f"B{row}:D{row}")

        headers = ["Subgroup", "ICER", "Interpretation"]
        for i, header in enumerate(headers):
            ws.cell(row=row+1, column=2+i, value=header).style = "header_style"

        if precomputed and "subgroups" in precomputed:
            for i, result in enumerate(precomputed["subgroups"]):
                r = row + 2 + i
                ws.cell(row=r, column=2, value=result["subgroup"])
                ws.cell(row=r, column=3, value=result["icer"]).number_format = '"$"#,##0'
                ws.cell(row=r, column=4, value=result["interpretation"])
        else:
            subgroups = ["Diabetic Patients", "CKD Stage 3+", "Prior CV Event", "Age > 70"]
            for i, sg in enumerate(subgroups):
                r = row + 2 + i
                ws.cell(row=r, column=2, value=sg)
                ws.cell(row=r, column=3, value="[Run scenarios]")
                ws.cell(row=r, column=4, value="-")

        # Set column widths
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20


def create_cea_template(output_path: str, precomputed: Optional[Dict] = None) -> str:
    """Create CEA Excel template."""
    template = CEAExcelTemplate()
    return template.generate(output_path, precomputed)
