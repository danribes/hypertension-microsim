"""
Streamlit Web Interface for Hypertension Microsimulation Model.

Cost-Effectiveness Analysis comparing IXA-001 vs Spironolactone
in adults with resistant hypertension.

Enhanced version with full parameter exposure.
"""

import streamlit as st
import numpy as np
import pandas as pd
from typing import Optional, Dict, List, Any
from io import BytesIO
import sys
from pathlib import Path
from dataclasses import dataclass, field

# Add src to path
sys.path.insert(0, str(Path(__file__).parent))

from src import run_cea, CEAResults, Treatment
from src.population import PopulationParams, PopulationGenerator
from src.patient import Patient, CardiacState, RenalState, NeuroState
from src.simulation import Simulation, SimulationConfig, SimulationResults
from src.risk_assessment import BaselineRiskProfile
from src.costs.costs import CostInputs, US_COSTS, UK_COSTS

# Page configuration
st.set_page_config(
    page_title="CEA Microsimulation - IXA-001",
    page_icon="🫀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2rem;
        font-weight: bold;
        color: #1F4E79;
        margin-bottom: 0.5rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        border-radius: 8px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .risk-high { color: #dc3545; font-weight: bold; }
    .risk-moderate { color: #fd7e14; }
    .risk-low { color: #28a745; }
    .param-section {
        border-left: 3px solid #1F4E79;
        padding-left: 10px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)


@dataclass
class ExtendedPopulationParams(PopulationParams):
    """Extended population parameters with additional comorbidities."""
    # Additional comorbidities
    copd_prev: float = 0.17
    depression_prev: float = 0.27
    anxiety_prev: float = 0.17
    substance_use_prev: float = 0.10
    smi_prev: float = 0.04
    afib_prev: float = 0.10
    pad_prev: float = 0.15

    # SGLT2i settings
    sglt2_uptake_rate: float = 0.40

    # Additional bounds
    age_min: float = 40.0
    age_max: float = 85.0
    sbp_min: float = 140.0
    sbp_max: float = 200.0
    egfr_min: float = 15.0
    egfr_max: float = 120.0


@dataclass
class TreatmentParams:
    """Configurable treatment effect parameters."""
    # IXA-001
    ixa_sbp_reduction: float = 20.0
    ixa_sbp_reduction_sd: float = 8.0
    ixa_discontinuation_rate: float = 0.12

    # Spironolactone
    spiro_sbp_reduction: float = 9.0
    spiro_sbp_reduction_sd: float = 6.0
    spiro_discontinuation_rate: float = 0.15

    # Standard care
    standard_sbp_reduction: float = 3.0
    standard_discontinuation_rate: float = 0.10

    # Adherence
    adherence_effect_multiplier: float = 0.30  # Effect when non-adherent


@dataclass
class ClinicalParams:
    """Configurable clinical parameters."""
    # Case fatality rates (30-day mortality)
    cfr_mi: float = 0.08
    cfr_ischemic_stroke: float = 0.10
    cfr_hemorrhagic_stroke: float = 0.25
    cfr_hf: float = 0.05

    # Stroke distribution
    stroke_ischemic_fraction: float = 0.85

    # Post-event mortality (annual)
    post_mi_year1_mortality: float = 0.05
    post_mi_thereafter_mortality: float = 0.03
    post_stroke_year1_mortality: float = 0.10
    post_stroke_thereafter_mortality: float = 0.05
    hf_annual_mortality: float = 0.08
    esrd_annual_mortality: float = 0.15

    # Prior event risk multipliers
    prior_mi_multiplier: float = 2.5
    prior_stroke_multiplier: float = 3.0
    prior_hf_multiplier: float = 2.0
    prior_tia_multiplier: float = 2.0

    # Cognitive decline rates (annual)
    normal_to_mci_rate: float = 0.02
    mci_to_dementia_rate: float = 0.10
    normal_to_dementia_rate: float = 0.005

    # eGFR decline (annual mL/min)
    egfr_decline_under_40: float = 0.0
    egfr_decline_40_65: float = 1.0
    egfr_decline_over_65: float = 1.5

    # Safety thresholds
    hyperkalemia_threshold: float = 5.5  # K+ mmol/L


@dataclass
class UtilityParams:
    """Configurable utility/QALY parameters."""
    # Baseline by age
    baseline_utility_40: float = 0.90
    baseline_utility_50: float = 0.87
    baseline_utility_60: float = 0.84
    baseline_utility_70: float = 0.80
    baseline_utility_80: float = 0.75
    baseline_utility_90: float = 0.70

    # Disutilities (decrements)
    disutility_uncontrolled_htn: float = 0.04
    disutility_post_mi: float = 0.12
    disutility_post_stroke: float = 0.18
    disutility_acute_hf: float = 0.25
    disutility_chronic_hf: float = 0.15
    disutility_ckd_3a: float = 0.01
    disutility_ckd_3b: float = 0.03
    disutility_ckd_4: float = 0.06
    disutility_esrd: float = 0.35
    disutility_diabetes: float = 0.04

    # Acute event disutilities
    acute_disutility_mi: float = 0.20
    acute_disutility_stroke: float = 0.40
    acute_disutility_hf: float = 0.25


@dataclass
class CustomCostInputs:
    """Custom cost inputs allowing user modification."""
    # Drug costs (monthly)
    ixa_001_monthly: float = 500.0
    spironolactone_monthly: float = 15.0
    sglt2_inhibitor_monthly: float = 450.0
    background_therapy_monthly: float = 75.0
    lab_test_cost_k: float = 15.0

    # Acute event costs
    mi_acute: float = 25000.0
    ischemic_stroke_acute: float = 15200.0
    hemorrhagic_stroke_acute: float = 22500.0
    tia_acute: float = 2100.0
    hf_admission: float = 18000.0

    # Annual management costs
    controlled_htn_annual: float = 800.0
    uncontrolled_htn_annual: float = 1200.0
    post_mi_annual: float = 5500.0
    post_stroke_annual: float = 12000.0
    heart_failure_annual: float = 15000.0
    ckd_stage_3a_annual: float = 2500.0
    ckd_stage_3b_annual: float = 4500.0
    ckd_stage_4_annual: float = 8000.0
    esrd_annual: float = 90000.0

    # Indirect costs
    daily_wage: float = 240.0
    absenteeism_mi_days: int = 7
    absenteeism_stroke_days: int = 30
    absenteeism_hf_days: int = 5
    disability_multiplier_stroke: float = 0.20
    disability_multiplier_hf: float = 0.15


def format_currency(value: float, symbol: str = "$") -> str:
    """Format currency value."""
    if abs(value) >= 1_000_000:
        return f"{symbol}{value/1_000_000:,.1f}M"
    elif abs(value) >= 1_000:
        return f"{symbol}{value/1_000:,.1f}K"
    else:
        return f"{symbol}{value:,.0f}"


def run_simulation_with_progress(
    n_patients: int,
    time_horizon: int,
    perspective: str,
    seed: int,
    discount_rate: float,
    pop_params: PopulationParams,
    status_container,
    custom_costs: Optional[CustomCostInputs] = None,
    treatment_params: Optional[TreatmentParams] = None,
    clinical_params: Optional[ClinicalParams] = None,
) -> tuple:
    """Run the CEA simulation with progress indicators."""

    # Update population params
    pop_params.n_patients = n_patients
    pop_params.seed = seed

    total_cycles = time_horizon * 12  # Monthly cycles

    # Create simulation config
    config = SimulationConfig(
        n_patients=n_patients,
        time_horizon_months=total_cycles,
        seed=seed,
        cost_perspective=perspective,
        discount_rate=discount_rate,
        show_progress=False
    )

    # Create progress elements inside the status container using write
    progress_placeholder = status_container.empty()

    def update_progress(phase: str, pct: int, detail: str):
        """Update progress display."""
        progress_placeholder.markdown(f"""
**{phase}**

{detail}

{'█' * (pct // 5)}{'░' * (20 - pct // 5)} {pct}%
""")

    # ===== Phase 1: Generate IXA-001 Population =====
    update_progress("Phase 1/5: Generating IXA-001 population", 0, "Creating patient cohort...")
    generator = PopulationGenerator(pop_params)
    patients_ixa = generator.generate()
    baseline_profiles_ixa = [p.baseline_risk_profile for p in patients_ixa]
    update_progress("Phase 1/5: Generating IXA-001 population", 100, f"Generated {n_patients} patients")

    # ===== Phase 2: Run IXA-001 Simulation =====
    sim = Simulation(config)

    # Apply custom costs if provided
    if custom_costs:
        _apply_custom_costs(sim, custom_costs)

    # Run simulation with progress updates
    results_ixa = _run_simulation_with_callback(
        sim, patients_ixa, Treatment.IXA_001, total_cycles,
        lambda pct, txt: update_progress("Phase 2/5: Simulating IXA-001 arm", pct, txt),
        "IXA-001", treatment_params, clinical_params
    )

    # ===== Phase 3: Generate Spironolactone Population =====
    update_progress("Phase 3/5: Generating Spironolactone population", 0, "Creating comparator cohort...")

    pop_params_comp = PopulationParams(
        n_patients=n_patients, seed=seed,
        age_mean=pop_params.age_mean, age_sd=pop_params.age_sd,
        prop_male=pop_params.prop_male,
        sbp_mean=pop_params.sbp_mean, sbp_sd=pop_params.sbp_sd,
        egfr_mean=pop_params.egfr_mean, egfr_sd=pop_params.egfr_sd,
        uacr_mean=pop_params.uacr_mean, uacr_sd=pop_params.uacr_sd,
        total_chol_mean=pop_params.total_chol_mean, hdl_chol_mean=pop_params.hdl_chol_mean,
        bmi_mean=pop_params.bmi_mean, bmi_sd=pop_params.bmi_sd,
        diabetes_prev=pop_params.diabetes_prev, smoker_prev=pop_params.smoker_prev,
        dyslipidemia_prev=pop_params.dyslipidemia_prev,
        prior_mi_prev=pop_params.prior_mi_prev, prior_stroke_prev=pop_params.prior_stroke_prev,
        heart_failure_prev=pop_params.heart_failure_prev,
        adherence_prob=pop_params.adherence_prob,
    )
    generator_comp = PopulationGenerator(pop_params_comp)
    patients_spi = generator_comp.generate()
    baseline_profiles_spi = [p.baseline_risk_profile for p in patients_spi]
    update_progress("Phase 3/5: Generating Spironolactone population", 100, f"Generated {n_patients} patients")

    # ===== Phase 4: Run Spironolactone Simulation =====
    results_spi = _run_simulation_with_callback(
        sim, patients_spi, Treatment.SPIRONOLACTONE, total_cycles,
        lambda pct, txt: update_progress("Phase 4/5: Simulating Spironolactone arm", pct, txt),
        "Spironolactone", treatment_params, clinical_params
    )

    # ===== Phase 5: Calculate Results =====
    update_progress("Phase 5/5: Calculating cost-effectiveness", 50, "Computing ICER and outcomes...")

    cea = CEAResults(intervention=results_ixa, comparator=results_spi)
    cea.calculate_icer()

    update_progress("Phase 5/5: Calculating cost-effectiveness", 100, "Analysis complete!")

    # Clear progress and show completion
    progress_placeholder.empty()
    status_container.update(label="Simulation complete!", state="complete")

    return cea, patients_ixa, patients_spi, baseline_profiles_ixa


def _apply_custom_costs(sim: Simulation, custom_costs: CustomCostInputs):
    """Apply custom cost parameters to simulation."""
    sim.costs.ixa_001_monthly = custom_costs.ixa_001_monthly
    sim.costs.spironolactone_monthly = custom_costs.spironolactone_monthly
    sim.costs.sglt2_inhibitor_monthly = custom_costs.sglt2_inhibitor_monthly
    sim.costs.background_therapy_monthly = custom_costs.background_therapy_monthly
    sim.costs.lab_test_cost_k = custom_costs.lab_test_cost_k
    sim.costs.mi_acute = custom_costs.mi_acute
    sim.costs.ischemic_stroke_acute = custom_costs.ischemic_stroke_acute
    sim.costs.hemorrhagic_stroke_acute = custom_costs.hemorrhagic_stroke_acute
    sim.costs.tia_acute = custom_costs.tia_acute
    sim.costs.hf_admission = custom_costs.hf_admission
    sim.costs.controlled_htn_annual = custom_costs.controlled_htn_annual
    sim.costs.uncontrolled_htn_annual = custom_costs.uncontrolled_htn_annual
    sim.costs.post_mi_annual = custom_costs.post_mi_annual
    sim.costs.post_stroke_annual = custom_costs.post_stroke_annual
    sim.costs.heart_failure_annual = custom_costs.heart_failure_annual
    sim.costs.ckd_stage_3a_annual = custom_costs.ckd_stage_3a_annual
    sim.costs.ckd_stage_3b_annual = custom_costs.ckd_stage_3b_annual
    sim.costs.ckd_stage_4_annual = custom_costs.ckd_stage_4_annual
    sim.costs.esrd_annual = custom_costs.esrd_annual
    sim.costs.daily_wage = custom_costs.daily_wage
    sim.costs.absenteeism_acute_mi_days = custom_costs.absenteeism_mi_days
    sim.costs.absenteeism_stroke_days = custom_costs.absenteeism_stroke_days
    sim.costs.absenteeism_hf_days = custom_costs.absenteeism_hf_days
    sim.costs.disability_multiplier_stroke = custom_costs.disability_multiplier_stroke
    sim.costs.disability_multiplier_hf = custom_costs.disability_multiplier_hf


def _run_simulation_with_callback(sim, patients, treatment, total_cycles, progress_callback, arm_name,
                                   treatment_params=None, clinical_params=None, n_sample_patients=5):
    """Run simulation with progress updates and detailed logging for sample patients."""
    from src.patient import Treatment as TreatmentEnum

    results = SimulationResults(treatment=treatment, n_patients=len(patients))

    # Initialize detailed simulation log for sample patients
    sample_ids = list(range(min(n_sample_patients, len(patients))))
    simulation_log = {pid: {
        'patient_id': pid,
        'initial_age': patients[pid].age,
        'initial_sbp': patients[pid].current_sbp,
        'initial_egfr': patients[pid].egfr,
        'treatment': treatment.value,
        'has_diabetes': patients[pid].has_diabetes,
        'has_hf': patients[pid].has_heart_failure,
        'cycles': []
    } for pid in sample_ids}

    # Assign treatment to all patients
    for patient in patients:
        sim.treatment_mgr.assign_treatment(patient, treatment)
        if patient.on_sglt2_inhibitor:
            results.sglt2_users += 1

    n_cycles = int(sim.config.time_horizon_months / sim.config.cycle_length_months)
    update_interval = max(1, n_cycles // 20)  # Update progress ~20 times
    log_interval = max(1, n_cycles // 60)  # Log ~60 times (every ~8 months for 40yr sim)

    for cycle in range(n_cycles):
        # Update progress bar periodically
        if cycle % update_interval == 0:
            progress_pct = int((cycle / n_cycles) * 100)
            years_simulated = cycle / 12
            progress_callback(progress_pct, f"Simulating {arm_name}: Year {years_simulated:.1f}/{sim.config.time_horizon_months/12:.0f}")

        for patient in patients:
            if not patient.is_alive:
                continue

            # Check adherence
            adherence_changed = sim.adherence_transition.check_adherence_change(patient)
            if adherence_changed:
                sim.treatment_mgr.update_effect_for_adherence(patient)

            # Safety checks for Spironolactone
            is_quarterly = (int(patient.time_in_simulation) % 3 == 0)
            hyperkalemia_stop = False
            if is_quarterly and patient.treatment == TreatmentEnum.SPIRONOLACTONE:
                patient.accrue_costs(sim.costs.lab_test_cost_k)
                if sim.treatment_mgr.check_safety_stop_rules(patient):
                    sim.treatment_mgr.assign_treatment(patient, TreatmentEnum.STANDARD_CARE)
                    patient.hyperkalemia_history += 1
                    hyperkalemia_stop = True

            # Neuro progression
            old_neuro = patient.neuro_state
            sim.neuro_transition.check_neuro_progression(patient)
            neuro_changed = patient.neuro_state != old_neuro
            if neuro_changed and patient.neuro_state.value == "dementia":
                results.dementia_cases += 1

            # Cardiac events - calculate transition probabilities
            probs = sim.transition_calc.calculate_transitions(patient)
            new_event = sim.transition_calc.sample_event(patient, probs)

            # Log detailed calculations for sample patients
            if patient.patient_id in sample_ids and cycle % log_interval == 0:
                cycle_log = {
                    'cycle': cycle,
                    'year': cycle / 12,
                    'age': patient.age,
                    'sbp': patient.current_sbp,
                    'true_sbp': getattr(patient, 'true_mean_sbp', patient.current_sbp),
                    'egfr': patient.egfr,
                    'is_adherent': patient.is_adherent,
                    'adherence_changed': adherence_changed,
                    'cardiac_state': patient.cardiac_state.value if hasattr(patient.cardiac_state, 'value') else str(patient.cardiac_state),
                    'renal_state': patient.renal_state.value if hasattr(patient.renal_state, 'value') else str(patient.renal_state),
                    'neuro_state': patient.neuro_state.value if hasattr(patient.neuro_state, 'value') else str(patient.neuro_state),
                    'probs': {
                        'p_mi': probs.to_mi,
                        'p_ischemic_stroke': probs.to_ischemic_stroke,
                        'p_hemorrhagic_stroke': probs.to_hemorrhagic_stroke,
                        'p_tia': probs.to_tia,
                        'p_hf': probs.to_hf,
                        'p_cv_death': probs.to_cv_death,
                        'p_non_cv_death': probs.to_non_cv_death,
                    },
                    'event': new_event.value if hasattr(new_event, 'value') else str(new_event) if new_event else None,
                    'neuro_changed': neuro_changed,
                    'hyperkalemia_stop': hyperkalemia_stop,
                    'cumulative_costs': patient.cumulative_costs,
                    'cumulative_qalys': patient.cumulative_qalys,
                    'treatment_effect': patient._treatment_effect_mmhg,
                }
                simulation_log[patient.patient_id]['cycles'].append(cycle_log)

            if new_event:
                if new_event == "NON_CV_DEATH":
                    results.non_cv_deaths += 1
                    patient.cardiac_state = "non_cv_death"
                else:
                    sim._record_event(new_event, results)
                    patient.transition_cardiac(new_event)

                    from src.costs.costs import get_event_cost, get_acute_absenteeism_cost
                    event_cost = get_event_cost(new_event.value, sim.costs)
                    absenteeism_cost = get_acute_absenteeism_cost(new_event.value, sim.costs, patient.age)

                    years = patient.time_in_simulation / 12
                    discount = 1 / ((1 + sim.config.discount_rate) ** years)

                    patient.accrue_costs(event_cost * discount)
                    results.total_costs += event_cost * discount
                    results.total_indirect_costs += absenteeism_cost * discount

            if not patient.is_alive:
                continue

            # Accrue outcomes
            sim._accrue_outcomes(patient, results)

            # Update SBP
            old_sbp = patient.current_sbp
            patient.update_sbp(patient._treatment_effect_mmhg, sim.rng)

            # Advance time and check renal
            old_renal = patient.renal_state
            patient.advance_time(sim.config.cycle_length_months)

            from src.patient import RenalState
            if patient.renal_state != old_renal:
                if patient.renal_state == RenalState.ESRD:
                    results.esrd_events += 1
                elif patient.renal_state == RenalState.CKD_STAGE_4:
                    results.ckd_4_events += 1

            # Check discontinuation
            if sim.treatment_mgr.check_discontinuation(patient):
                patient.treatment = TreatmentEnum.STANDARD_CARE

    # Final progress update
    progress_callback(100, f"{arm_name} simulation complete!")

    # Store patient results
    for patient in patients:
        results.patient_results.append(patient.to_dict())

    results.calculate_means()

    # Attach simulation log to results
    results.simulation_log = simulation_log

    return results


def analyze_subgroups(patients: List[Patient], results: SimulationResults, profiles: List[BaselineRiskProfile]) -> Dict:
    """Analyze results by subgroups."""
    subgroup_data = {
        'framingham': {'Low': [], 'Borderline': [], 'Intermediate': [], 'High': []},
        'kdigo': {'Low': [], 'Moderate': [], 'High': [], 'Very High': []},
        'gcua': {'I': [], 'II': [], 'III': [], 'IV': [], 'Moderate': [], 'Low': []},
        'age': {'<60': [], '60-70': [], '70-80': [], '80+': []},
        'ckd_stage': {'Stage 1-2': [], 'Stage 3a': [], 'Stage 3b': [], 'Stage 4': [], 'ESRD': []},
    }

    for i, (patient, profile) in enumerate(zip(patients, profiles)):
        patient_data = results.patient_results[i] if i < len(results.patient_results) else {}

        # Framingham category
        if profile.framingham_category:
            if profile.framingham_category in subgroup_data['framingham']:
                subgroup_data['framingham'][profile.framingham_category].append(patient_data)

        # KDIGO risk level
        if profile.kdigo_risk_level:
            if profile.kdigo_risk_level in subgroup_data['kdigo']:
                subgroup_data['kdigo'][profile.kdigo_risk_level].append(patient_data)

        # GCUA phenotype
        if profile.gcua_phenotype:
            if profile.gcua_phenotype in subgroup_data['gcua']:
                subgroup_data['gcua'][profile.gcua_phenotype].append(patient_data)

        # Age groups
        age = patient_data.get('age', patient.age)
        if age < 60:
            subgroup_data['age']['<60'].append(patient_data)
        elif age < 70:
            subgroup_data['age']['60-70'].append(patient_data)
        elif age < 80:
            subgroup_data['age']['70-80'].append(patient_data)
        else:
            subgroup_data['age']['80+'].append(patient_data)

        # CKD stage
        renal = patient_data.get('renal_state', patient.renal_state.value if hasattr(patient.renal_state, 'value') else str(patient.renal_state))
        if 'stage_1_2' in str(renal).lower() or 'ckd_stage_1_2' in str(renal).lower():
            subgroup_data['ckd_stage']['Stage 1-2'].append(patient_data)
        elif 'stage_3a' in str(renal).lower():
            subgroup_data['ckd_stage']['Stage 3a'].append(patient_data)
        elif 'stage_3b' in str(renal).lower():
            subgroup_data['ckd_stage']['Stage 3b'].append(patient_data)
        elif 'stage_4' in str(renal).lower():
            subgroup_data['ckd_stage']['Stage 4'].append(patient_data)
        elif 'esrd' in str(renal).lower():
            subgroup_data['ckd_stage']['ESRD'].append(patient_data)

    return subgroup_data


def generate_excel_report(cea: CEAResults, pop_params: PopulationParams,
                          subgroup_data: Dict, currency: str,
                          custom_costs: Optional[CustomCostInputs] = None,
                          treatment_params: Optional[TreatmentParams] = None,
                          clinical_params: Optional[ClinicalParams] = None) -> BytesIO:
    """Generate comprehensive Excel report with charts and formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference

    wb = Workbook()

    # ===== Style definitions =====
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    alt_row_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    title_font = Font(bold=True, size=18, color="1F4E79")
    subtitle_font = Font(bold=True, size=14, color="2E75B6")
    border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    currency_sym = currency

    def apply_table_style(ws, start_row, end_row, num_cols):
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if row_idx == start_row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                elif (row_idx - start_row) % 2 == 0:
                    cell.fill = alt_row_fill

    # ========== Sheet 1: Executive Summary ==========
    ws = wb.active
    ws.title = "Executive Summary"

    ws.merge_cells('A1:D1')
    ws['A1'] = "Cost-Effectiveness Analysis Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = "IXA-001 vs Spironolactone in Resistant Hypertension"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Key Results
    ws['A4'] = "KEY RESULTS"
    ws['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A4'].fill = header_fill
    ws.merge_cells('A4:B4')

    results_data = [
        ["Incremental Costs", f"{currency_sym}{cea.incremental_costs:,.0f}"],
        ["Incremental QALYs", f"{cea.incremental_qalys:.3f}"],
        ["ICER", f"{currency_sym}{cea.icer:,.0f}/QALY" if cea.icer else "Dominant"],
        ["Interpretation", "Cost-Effective" if (cea.icer and cea.icer < 100000) or cea.incremental_qalys > 0 else "Review Required"],
    ]

    for i, (label, value) in enumerate(results_data, start=5):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
        if "Dominant" in str(value) or "Cost-Effective" in str(value):
            ws.cell(row=i, column=2).fill = highlight_fill

    # Population Summary
    ws['A10'] = "POPULATION CHARACTERISTICS"
    ws['A10'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A10'].fill = header_fill
    ws.merge_cells('A10:B10')

    pop_data = [
        ["Cohort Size", f"{pop_params.n_patients:,} per arm"],
        ["Mean Age", f"{pop_params.age_mean:.0f} years (SD {pop_params.age_sd:.0f})"],
        ["% Male", f"{pop_params.prop_male*100:.0f}%"],
        ["Mean SBP", f"{pop_params.sbp_mean:.0f} mmHg"],
        ["Mean eGFR", f"{pop_params.egfr_mean:.0f} mL/min/1.73m²"],
        ["Diabetes", f"{pop_params.diabetes_prev*100:.0f}%"],
        ["Prior MI", f"{pop_params.prior_mi_prev*100:.0f}%"],
        ["Heart Failure", f"{pop_params.heart_failure_prev*100:.0f}%"],
    ]

    for i, (label, value) in enumerate(pop_data, start=11):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = alt_row_fill
            ws.cell(row=i, column=2).fill = alt_row_fill

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # ========== Sheet 2: Clinical Events ==========
    ws2 = wb.create_sheet("Clinical Events")

    ws2['A1'] = "Clinical Events Comparison"
    ws2['A1'].font = title_font

    events_header = ["Event", "IXA-001", "Spironolactone", "Difference"]
    events_data = [
        ["MI", cea.intervention.mi_events, cea.comparator.mi_events],
        ["Stroke (Total)", cea.intervention.stroke_events, cea.comparator.stroke_events],
        ["  Ischemic", cea.intervention.ischemic_stroke_events, cea.comparator.ischemic_stroke_events],
        ["  Hemorrhagic", cea.intervention.hemorrhagic_stroke_events, cea.comparator.hemorrhagic_stroke_events],
        ["TIA", cea.intervention.tia_events, cea.comparator.tia_events],
        ["Heart Failure", cea.intervention.hf_events, cea.comparator.hf_events],
        ["CV Death", cea.intervention.cv_deaths, cea.comparator.cv_deaths],
        ["Non-CV Death", cea.intervention.non_cv_deaths, cea.comparator.non_cv_deaths],
        ["CKD Stage 4", cea.intervention.ckd_4_events, cea.comparator.ckd_4_events],
        ["ESRD", cea.intervention.esrd_events, cea.comparator.esrd_events],
        ["Dementia", cea.intervention.dementia_cases, cea.comparator.dementia_cases],
    ]

    ws2.append([])
    ws2.append(events_header)
    for row in events_data:
        diff = row[2] - row[1]  # Comparator - Intervention (positive = events avoided)
        ws2.append([row[0], row[1], row[2], diff])

    apply_table_style(ws2, 3, 3 + len(events_data), 4)

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Clinical Events per 1000 Patients"
    chart.y_axis.title = "Number of Events"
    chart.style = 10

    data = Reference(ws2, min_col=2, min_row=3, max_col=3, max_row=9)
    cats = Reference(ws2, min_col=1, min_row=4, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10

    ws2.add_chart(chart, "F3")

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15

    # ========== Sheet 3: Cost Analysis ==========
    ws3 = wb.create_sheet("Cost Analysis")

    ws3['A1'] = "Cost Breakdown Analysis"
    ws3['A1'].font = title_font

    direct_ixa = cea.intervention.mean_costs - (cea.intervention.total_indirect_costs / cea.intervention.n_patients)
    indirect_ixa = cea.intervention.total_indirect_costs / cea.intervention.n_patients
    direct_spi = cea.comparator.mean_costs - (cea.comparator.total_indirect_costs / cea.comparator.n_patients)
    indirect_spi = cea.comparator.total_indirect_costs / cea.comparator.n_patients

    cost_header = ["Cost Category", "IXA-001", "Spironolactone", "Difference"]
    cost_data = [
        ["Direct Costs", f"{currency_sym}{direct_ixa:,.0f}", f"{currency_sym}{direct_spi:,.0f}", f"{currency_sym}{direct_ixa - direct_spi:,.0f}"],
        ["Indirect Costs", f"{currency_sym}{indirect_ixa:,.0f}", f"{currency_sym}{indirect_spi:,.0f}", f"{currency_sym}{indirect_ixa - indirect_spi:,.0f}"],
        ["Total Costs", f"{currency_sym}{cea.intervention.mean_costs:,.0f}", f"{currency_sym}{cea.comparator.mean_costs:,.0f}", f"{currency_sym}{cea.incremental_costs:,.0f}"],
    ]

    ws3.append([])
    ws3.append(cost_header)
    for row in cost_data:
        ws3.append(row)

    apply_table_style(ws3, 3, 3 + len(cost_data), 4)

    # Cost parameters used (if custom)
    if custom_costs:
        ws3['A10'] = "COST PARAMETERS USED"
        ws3['A10'].font = subtitle_font

        cost_params = [
            ["Parameter", "Value"],
            ["IXA-001 Monthly", f"{currency_sym}{custom_costs.ixa_001_monthly:,.0f}"],
            ["Spironolactone Monthly", f"{currency_sym}{custom_costs.spironolactone_monthly:,.0f}"],
            ["SGLT2i Monthly", f"{currency_sym}{custom_costs.sglt2_inhibitor_monthly:,.0f}"],
            ["MI Acute", f"{currency_sym}{custom_costs.mi_acute:,.0f}"],
            ["Stroke (Ischemic)", f"{currency_sym}{custom_costs.ischemic_stroke_acute:,.0f}"],
            ["Stroke (Hemorrhagic)", f"{currency_sym}{custom_costs.hemorrhagic_stroke_acute:,.0f}"],
            ["HF Admission", f"{currency_sym}{custom_costs.hf_admission:,.0f}"],
            ["ESRD Annual", f"{currency_sym}{custom_costs.esrd_annual:,.0f}"],
        ]

        for i, row in enumerate(cost_params, start=11):
            for j, val in enumerate(row, start=1):
                cell = ws3.cell(row=i, column=j, value=val)
                if i == 11:
                    cell.font = header_font
                    cell.fill = subheader_fill

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 20

    # ========== Sheet 4: Subgroup Analysis ==========
    ws4 = wb.create_sheet("Subgroup Analysis")

    ws4['A1'] = "Subgroup Analysis"
    ws4['A1'].font = title_font

    current_row = 3

    for subgroup_name, subgroup_cats in [
        ("Age Group", subgroup_data['age']),
        ("Framingham CVD Risk", subgroup_data['framingham']),
        ("KDIGO Risk Level", subgroup_data['kdigo']),
        ("GCUA Phenotype", subgroup_data['gcua']),
    ]:
        ws4.cell(row=current_row, column=1, value=f"By {subgroup_name}").font = subtitle_font
        current_row += 1

        ws4.cell(row=current_row, column=1, value="Category")
        ws4.cell(row=current_row, column=2, value="N Patients")
        ws4.cell(row=current_row, column=3, value="Mean Costs")
        ws4.cell(row=current_row, column=4, value="Mean QALYs")
        header_row = current_row

        for cat, patients in subgroup_cats.items():
            n = len(patients)
            if n > 0:
                current_row += 1
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                ws4.cell(row=current_row, column=1, value=cat)
                ws4.cell(row=current_row, column=2, value=n)
                ws4.cell(row=current_row, column=3, value=f"{currency_sym}{mean_costs:,.0f}")
                ws4.cell(row=current_row, column=4, value=f"{mean_qalys:.3f}")

        apply_table_style(ws4, header_row, current_row, 4)
        current_row += 2

    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15

    # ========== Sheet 5: WTP Analysis ==========
    ws5 = wb.create_sheet("WTP Analysis")

    ws5['A1'] = "Willingness-to-Pay Analysis"
    ws5['A1'].font = title_font

    ws5.append([])
    wtp_header = ["WTP Threshold", "NMB IXA-001", "NMB Spironolactone", "Incremental NMB", "Cost-Effective?"]
    ws5.append(wtp_header)

    wtp_values = [0, 25000, 50000, 75000, 100000, 150000, 200000]

    for wtp in wtp_values:
        nmb_ixa = cea.intervention.mean_qalys * wtp - cea.intervention.mean_costs
        nmb_spi = cea.comparator.mean_qalys * wtp - cea.comparator.mean_costs
        inc_nmb = nmb_ixa - nmb_spi
        ce = "Yes" if inc_nmb > 0 else "No"
        ws5.append([f"{currency_sym}{wtp:,}/QALY", f"{currency_sym}{nmb_ixa:,.0f}", f"{currency_sym}{nmb_spi:,.0f}", f"{currency_sym}{inc_nmb:,.0f}", ce])

    apply_table_style(ws5, 3, 3 + len(wtp_values), 5)

    # Highlight cost-effective rows
    for row in range(4, 4 + len(wtp_values)):
        if ws5.cell(row=row, column=5).value == "Yes":
            for col in range(1, 6):
                ws5.cell(row=row, column=col).fill = highlight_fill

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws5.column_dimensions[col].width = 18

    # ========== Sheet 6: Parameters ==========
    ws6 = wb.create_sheet("Parameters")

    ws6['A1'] = "Simulation Parameters"
    ws6['A1'].font = title_font

    current_row = 3

    # Demographics
    ws6.cell(row=current_row, column=1, value="DEMOGRAPHICS").font = Font(bold=True, color="FFFFFF")
    ws6.cell(row=current_row, column=1).fill = header_fill
    ws6.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    demo_params = [
        ["Mean Age (years)", f"{pop_params.age_mean:.0f} (SD {pop_params.age_sd:.0f})"],
        ["Age Range", f"{pop_params.age_min:.0f} - {pop_params.age_max:.0f}"],
        ["% Male", f"{pop_params.prop_male*100:.0f}%"],
        ["Mean BMI", f"{pop_params.bmi_mean:.1f}"],
    ]

    for label, value in demo_params:
        ws6.cell(row=current_row, column=1, value=label)
        ws6.cell(row=current_row, column=2, value=value)
        current_row += 1

    current_row += 1

    # Clinical Parameters
    ws6.cell(row=current_row, column=1, value="CLINICAL PARAMETERS").font = Font(bold=True, color="FFFFFF")
    ws6.cell(row=current_row, column=1).fill = header_fill
    ws6.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    clinical_display = [
        ["Mean SBP (mmHg)", f"{pop_params.sbp_mean:.0f} (SD {pop_params.sbp_sd:.0f})"],
        ["Mean eGFR (mL/min)", f"{pop_params.egfr_mean:.0f} (SD {pop_params.egfr_sd:.0f})"],
        ["Mean UACR (mg/g)", f"{pop_params.uacr_mean:.0f}"],
    ]

    for label, value in clinical_display:
        ws6.cell(row=current_row, column=1, value=label)
        ws6.cell(row=current_row, column=2, value=value)
        current_row += 1

    current_row += 1

    # Comorbidities
    ws6.cell(row=current_row, column=1, value="COMORBIDITIES (%)").font = Font(bold=True, color="FFFFFF")
    ws6.cell(row=current_row, column=1).fill = header_fill
    ws6.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1

    comorbid_params = [
        ["Diabetes", f"{pop_params.diabetes_prev*100:.0f}%"],
        ["Current Smoker", f"{pop_params.smoker_prev*100:.0f}%"],
        ["Dyslipidemia", f"{pop_params.dyslipidemia_prev*100:.0f}%"],
        ["Prior MI", f"{pop_params.prior_mi_prev*100:.0f}%"],
        ["Prior Stroke", f"{pop_params.prior_stroke_prev*100:.0f}%"],
        ["Heart Failure", f"{pop_params.heart_failure_prev*100:.0f}%"],
    ]

    for label, value in comorbid_params:
        ws6.cell(row=current_row, column=1, value=label)
        ws6.cell(row=current_row, column=2, value=value)
        current_row += 1

    # Treatment params if provided
    if treatment_params:
        current_row += 1
        ws6.cell(row=current_row, column=1, value="TREATMENT EFFECTS").font = Font(bold=True, color="FFFFFF")
        ws6.cell(row=current_row, column=1).fill = header_fill
        ws6.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        treatment_display = [
            ["IXA-001 SBP Reduction", f"{treatment_params.ixa_sbp_reduction:.0f} mmHg (SD {treatment_params.ixa_sbp_reduction_sd:.0f})"],
            ["IXA-001 Discontinuation", f"{treatment_params.ixa_discontinuation_rate*100:.0f}%/year"],
            ["Spiro SBP Reduction", f"{treatment_params.spiro_sbp_reduction:.0f} mmHg (SD {treatment_params.spiro_sbp_reduction_sd:.0f})"],
            ["Spiro Discontinuation", f"{treatment_params.spiro_discontinuation_rate*100:.0f}%/year"],
        ]

        for label, value in treatment_display:
            ws6.cell(row=current_row, column=1, value=label)
            ws6.cell(row=current_row, column=2, value=value)
            current_row += 1

    # Clinical params if provided
    if clinical_params:
        current_row += 1
        ws6.cell(row=current_row, column=1, value="CLINICAL MODEL PARAMETERS").font = Font(bold=True, color="FFFFFF")
        ws6.cell(row=current_row, column=1).fill = header_fill
        ws6.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        model_display = [
            ["MI Case Fatality", f"{clinical_params.cfr_mi*100:.0f}%"],
            ["Ischemic Stroke CFR", f"{clinical_params.cfr_ischemic_stroke*100:.0f}%"],
            ["Hemorrhagic Stroke CFR", f"{clinical_params.cfr_hemorrhagic_stroke*100:.0f}%"],
            ["HF Case Fatality", f"{clinical_params.cfr_hf*100:.0f}%"],
            ["Stroke Ischemic %", f"{clinical_params.stroke_ischemic_fraction*100:.0f}%"],
        ]

        for label, value in model_display:
            ws6.cell(row=current_row, column=1, value=label)
            ws6.cell(row=current_row, column=2, value=value)
            current_row += 1

    ws6.column_dimensions['A'].width = 30
    ws6.column_dimensions['B'].width = 25

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============== DISPLAY FUNCTIONS ==============

def display_key_metrics(cea: CEAResults, currency: str):
    """Display key CEA metrics."""
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(label="Incremental Costs", value=format_currency(cea.incremental_costs, currency))

    with col2:
        st.metric(label="Incremental QALYs", value=f"{cea.incremental_qalys:.3f}")

    with col3:
        if cea.icer is not None:
            st.metric(label="ICER", value=f"{currency}{cea.icer:,.0f}/QALY")
        else:
            st.metric(label="ICER", value="Dominant" if cea.incremental_qalys > 0 else "Dominated")

    with col4:
        if cea.icer is not None:
            if cea.icer < 50000:
                interpretation = "Highly Cost-Effective"
            elif cea.icer < 100000:
                interpretation = "Cost-Effective"
            elif cea.icer < 150000:
                interpretation = "Borderline"
            else:
                interpretation = "Not Cost-Effective"
        else:
            interpretation = "Dominant" if cea.incremental_qalys > 0 else "Dominated"

        st.metric(label="Interpretation", value=interpretation)


def display_outcomes_table(cea: CEAResults, currency: str):
    """Display outcomes comparison table."""
    st.markdown("### Clinical Outcomes Comparison")

    data = {
        "Outcome": [
            "Mean Total Costs", "Mean QALYs", "Mean Life Years",
            "MI Events", "Stroke Events", "  - Ischemic", "  - Hemorrhagic",
            "TIA Events", "Heart Failure", "CV Deaths", "Non-CV Deaths",
            "CKD Stage 4", "ESRD Events", "Dementia Cases",
        ],
        "IXA-001": [
            f"{currency}{cea.intervention.mean_costs:,.0f}",
            f"{cea.intervention.mean_qalys:.3f}",
            f"{cea.intervention.mean_life_years:.2f}",
            cea.intervention.mi_events, cea.intervention.stroke_events,
            cea.intervention.ischemic_stroke_events, cea.intervention.hemorrhagic_stroke_events,
            cea.intervention.tia_events, cea.intervention.hf_events,
            cea.intervention.cv_deaths, cea.intervention.non_cv_deaths,
            cea.intervention.ckd_4_events, cea.intervention.esrd_events,
            cea.intervention.dementia_cases,
        ],
        "Spironolactone": [
            f"{currency}{cea.comparator.mean_costs:,.0f}",
            f"{cea.comparator.mean_qalys:.3f}",
            f"{cea.comparator.mean_life_years:.2f}",
            cea.comparator.mi_events, cea.comparator.stroke_events,
            cea.comparator.ischemic_stroke_events, cea.comparator.hemorrhagic_stroke_events,
            cea.comparator.tia_events, cea.comparator.hf_events,
            cea.comparator.cv_deaths, cea.comparator.non_cv_deaths,
            cea.comparator.ckd_4_events, cea.comparator.esrd_events,
            cea.comparator.dementia_cases,
        ],
        "Difference": [
            f"{currency}{cea.incremental_costs:,.0f}",
            f"{cea.incremental_qalys:.3f}",
            f"{cea.intervention.mean_life_years - cea.comparator.mean_life_years:.2f}",
            cea.intervention.mi_events - cea.comparator.mi_events,
            cea.intervention.stroke_events - cea.comparator.stroke_events,
            cea.intervention.ischemic_stroke_events - cea.comparator.ischemic_stroke_events,
            cea.intervention.hemorrhagic_stroke_events - cea.comparator.hemorrhagic_stroke_events,
            cea.intervention.tia_events - cea.comparator.tia_events,
            cea.intervention.hf_events - cea.comparator.hf_events,
            cea.intervention.cv_deaths - cea.comparator.cv_deaths,
            cea.intervention.non_cv_deaths - cea.comparator.non_cv_deaths,
            cea.intervention.ckd_4_events - cea.comparator.ckd_4_events,
            cea.intervention.esrd_events - cea.comparator.esrd_events,
            cea.intervention.dementia_cases - cea.comparator.dementia_cases,
        ],
    }

    df = pd.DataFrame(data)
    st.dataframe(df, use_container_width=True, hide_index=True)


def display_detailed_costs(cea: CEAResults, currency: str):
    """Display detailed cost breakdown."""
    st.markdown("### Cost Breakdown")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**IXA-001**")
        direct_ixa = cea.intervention.mean_costs - (cea.intervention.total_indirect_costs / cea.intervention.n_patients)
        indirect_ixa = cea.intervention.total_indirect_costs / cea.intervention.n_patients

        cost_data_ixa = pd.DataFrame({
            'Category': ['Direct Costs', 'Indirect Costs (Productivity)', 'Total'],
            'Amount': [f"{currency}{direct_ixa:,.0f}", f"{currency}{indirect_ixa:,.0f}",
                      f"{currency}{cea.intervention.mean_costs:,.0f}"]
        })
        st.dataframe(cost_data_ixa, hide_index=True, use_container_width=True)

    with col2:
        st.markdown("**Spironolactone**")
        direct_spi = cea.comparator.mean_costs - (cea.comparator.total_indirect_costs / cea.comparator.n_patients)
        indirect_spi = cea.comparator.total_indirect_costs / cea.comparator.n_patients

        cost_data_spi = pd.DataFrame({
            'Category': ['Direct Costs', 'Indirect Costs (Productivity)', 'Total'],
            'Amount': [f"{currency}{direct_spi:,.0f}", f"{currency}{indirect_spi:,.0f}",
                      f"{currency}{cea.comparator.mean_costs:,.0f}"]
        })
        st.dataframe(cost_data_spi, hide_index=True, use_container_width=True)


def display_medication_adherence(cea: CEAResults):
    """Display medication and adherence metrics."""
    st.markdown("### Medication & BP Control")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("SGLT2 Users (IXA-001)", cea.intervention.sglt2_users)
    with col2:
        st.metric("SGLT2 Users (Spiro)", cea.comparator.sglt2_users)
    with col3:
        pct_controlled_ixa = (cea.intervention.time_controlled /
                              (cea.intervention.time_controlled + cea.intervention.time_uncontrolled) * 100
                              if (cea.intervention.time_controlled + cea.intervention.time_uncontrolled) > 0 else 0)
        st.metric("% Time BP Controlled (IXA)", f"{pct_controlled_ixa:.1f}%")
    with col4:
        pct_controlled_spi = (cea.comparator.time_controlled /
                              (cea.comparator.time_controlled + cea.comparator.time_uncontrolled) * 100
                              if (cea.comparator.time_controlled + cea.comparator.time_uncontrolled) > 0 else 0)
        st.metric("% Time BP Controlled (Spiro)", f"{pct_controlled_spi:.1f}%")


def display_subgroup_analysis(subgroup_data: Dict, currency: str):
    """Display subgroup analysis results."""
    st.markdown("### Subgroup Analysis")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Framingham Risk", "KDIGO Risk", "GCUA Phenotype", "Age Group", "CKD Stage"])

    with tab1:
        st.markdown("**By Framingham CVD Risk Category**")
        fram_data = []
        for cat in ['Low', 'Borderline', 'Intermediate', 'High']:
            patients = subgroup_data['framingham'][cat]
            n = len(patients)
            if n > 0:
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                fram_data.append({'Category': cat, 'N': n, 'Mean Costs': f"{currency}{mean_costs:,.0f}", 'Mean QALYs': f"{mean_qalys:.3f}"})
        if fram_data:
            st.dataframe(pd.DataFrame(fram_data), hide_index=True, use_container_width=True)
        else:
            st.info("No Framingham risk data available")

    with tab2:
        st.markdown("**By KDIGO Risk Level**")
        kdigo_data = []
        for cat in ['Low', 'Moderate', 'High', 'Very High']:
            patients = subgroup_data['kdigo'][cat]
            n = len(patients)
            if n > 0:
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                kdigo_data.append({'Level': cat, 'N': n, 'Mean Costs': f"{currency}{mean_costs:,.0f}", 'Mean QALYs': f"{mean_qalys:.3f}"})
        if kdigo_data:
            st.dataframe(pd.DataFrame(kdigo_data), hide_index=True, use_container_width=True)
        else:
            st.info("No KDIGO risk data available")

    with tab3:
        st.markdown("**By GCUA Phenotype** (Age 60+, eGFR >60)")
        gcua_data = []
        phenotype_names = {'I': 'Accelerated Ager', 'II': 'Silent Renal', 'III': 'Vascular Dominant', 'IV': 'Senescent', 'Moderate': 'Moderate', 'Low': 'Low'}
        for cat in ['I', 'II', 'III', 'IV', 'Moderate', 'Low']:
            patients = subgroup_data['gcua'][cat]
            n = len(patients)
            if n > 0:
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                gcua_data.append({'Phenotype': f"{cat} ({phenotype_names.get(cat, '')})", 'N': n, 'Mean Costs': f"{currency}{mean_costs:,.0f}", 'Mean QALYs': f"{mean_qalys:.3f}"})
        if gcua_data:
            st.dataframe(pd.DataFrame(gcua_data), hide_index=True, use_container_width=True)
        else:
            st.info("No GCUA phenotype data (requires age 60+, eGFR >60)")

    with tab4:
        st.markdown("**By Age Group**")
        age_data = []
        for cat in ['<60', '60-70', '70-80', '80+']:
            patients = subgroup_data['age'][cat]
            n = len(patients)
            if n > 0:
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                age_data.append({'Age': cat, 'N': n, 'Mean Costs': f"{currency}{mean_costs:,.0f}", 'Mean QALYs': f"{mean_qalys:.3f}"})
        if age_data:
            st.dataframe(pd.DataFrame(age_data), hide_index=True, use_container_width=True)

    with tab5:
        st.markdown("**By CKD Stage**")
        ckd_data = []
        for cat in ['Stage 1-2', 'Stage 3a', 'Stage 3b', 'Stage 4', 'ESRD']:
            patients = subgroup_data['ckd_stage'][cat]
            n = len(patients)
            if n > 0:
                mean_costs = np.mean([p.get('cumulative_costs', 0) for p in patients])
                mean_qalys = np.mean([p.get('cumulative_qalys', 0) for p in patients])
                ckd_data.append({'Stage': cat, 'N': n, 'Mean Costs': f"{currency}{mean_costs:,.0f}", 'Mean QALYs': f"{mean_qalys:.3f}"})
        if ckd_data:
            st.dataframe(pd.DataFrame(ckd_data), hide_index=True, use_container_width=True)


def display_risk_stratification(profiles: List[BaselineRiskProfile]):
    """Display baseline risk stratification summary."""
    st.markdown("### Baseline Risk Stratification")

    framingham_counts = {'Low': 0, 'Borderline': 0, 'Intermediate': 0, 'High': 0}
    kdigo_counts = {'Low': 0, 'Moderate': 0, 'High': 0, 'Very High': 0}
    gcua_counts = {'I': 0, 'II': 0, 'III': 0, 'IV': 0, 'Moderate': 0, 'Low': 0}

    for p in profiles:
        if p.framingham_category in framingham_counts:
            framingham_counts[p.framingham_category] += 1
        if p.kdigo_risk_level in kdigo_counts:
            kdigo_counts[p.kdigo_risk_level] += 1
        if p.gcua_phenotype in gcua_counts:
            gcua_counts[p.gcua_phenotype] += 1

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**Framingham CVD Risk**")
        fram_df = pd.DataFrame([{'Category': k, 'N': v, '%': f"{v/len(profiles)*100:.1f}%"} for k, v in framingham_counts.items() if v > 0])
        if not fram_df.empty:
            st.dataframe(fram_df, hide_index=True)

    with col2:
        st.markdown("**KDIGO Risk Level**")
        kdigo_df = pd.DataFrame([{'Level': k, 'N': v, '%': f"{v/len(profiles)*100:.1f}%"} for k, v in kdigo_counts.items() if v > 0])
        if not kdigo_df.empty:
            st.dataframe(kdigo_df, hide_index=True)

    with col3:
        st.markdown("**GCUA Phenotype**")
        gcua_df = pd.DataFrame([{'Phenotype': k, 'N': v, '%': f"{v/len(profiles)*100:.1f}%"} for k, v in gcua_counts.items() if v > 0])
        if not gcua_df.empty:
            st.dataframe(gcua_df, hide_index=True)
        else:
            st.info("GCUA applies to age 60+, eGFR >60")


def display_event_charts(cea: CEAResults):
    """Display event comparison charts."""
    st.markdown("### Event Comparison Charts")

    col1, col2 = st.columns(2)

    with col1:
        cardiac_data = pd.DataFrame({
            "Event": ["MI", "Stroke", "TIA", "HF", "CV Death"],
            "IXA-001": [cea.intervention.mi_events, cea.intervention.stroke_events, cea.intervention.tia_events, cea.intervention.hf_events, cea.intervention.cv_deaths],
            "Spironolactone": [cea.comparator.mi_events, cea.comparator.stroke_events, cea.comparator.tia_events, cea.comparator.hf_events, cea.comparator.cv_deaths],
        })
        st.bar_chart(cardiac_data.set_index("Event"), use_container_width=True)
        st.caption("Cardiac Events (per 1000 patients)")

    with col2:
        renal_data = pd.DataFrame({
            "Event": ["CKD Stage 4", "ESRD", "Dementia"],
            "IXA-001": [cea.intervention.ckd_4_events, cea.intervention.esrd_events, cea.intervention.dementia_cases],
            "Spironolactone": [cea.comparator.ckd_4_events, cea.comparator.esrd_events, cea.comparator.dementia_cases],
        })
        st.bar_chart(renal_data.set_index("Event"), use_container_width=True)
        st.caption("Renal & Neuro Events (per 1000 patients)")


def display_ce_plane(cea: CEAResults, currency: str):
    """Display cost-effectiveness plane."""
    st.markdown("### Cost-Effectiveness Plane")

    chart_data = pd.DataFrame({"Incremental QALYs": [cea.incremental_qalys], "Incremental Costs": [cea.incremental_costs]})
    st.scatter_chart(chart_data, x="Incremental QALYs", y="Incremental Costs", use_container_width=True)

    if cea.incremental_qalys > 0 and cea.incremental_costs < 0:
        st.success("**Quadrant: Southeast (Dominant)** - IXA-001 is more effective and less costly")
    elif cea.incremental_qalys > 0 and cea.incremental_costs > 0:
        st.info(f"**Quadrant: Northeast** - IXA-001 is more effective but more costly (ICER: {currency}{cea.icer:,.0f}/QALY)")
    elif cea.incremental_qalys < 0 and cea.incremental_costs < 0:
        st.warning("**Quadrant: Southwest** - IXA-001 is less effective but less costly")
    else:
        st.error("**Quadrant: Northwest (Dominated)** - IXA-001 is less effective and more costly")


def display_wtp_analysis(cea: CEAResults, currency: str):
    """Display willingness-to-pay analysis."""
    st.markdown("### Willingness-to-Pay Analysis")

    wtp_thresholds = [0, 25000, 50000, 75000, 100000, 150000, 200000]

    data = []
    for wtp in wtp_thresholds:
        nmb_ixa = cea.intervention.mean_qalys * wtp - cea.intervention.mean_costs
        nmb_spi = cea.comparator.mean_qalys * wtp - cea.comparator.mean_costs
        incremental_nmb = nmb_ixa - nmb_spi
        ce = "Yes" if incremental_nmb > 0 else "No"

        data.append({
            "WTP Threshold": f"{currency}{wtp:,}/QALY",
            "NMB IXA-001": f"{currency}{nmb_ixa:,.0f}",
            "NMB Spironolactone": f"{currency}{nmb_spi:,.0f}",
            "Incremental NMB": f"{currency}{incremental_nmb:,.0f}",
            "Cost-Effective?": ce,
        })

    df = pd.DataFrame(data)
    st.dataframe(df, use_container_width=True, hide_index=True)


def display_patient_trajectories(patients: List[Patient], results: SimulationResults):
    """Display individual patient trajectory analysis."""
    st.markdown("### Patient Trajectories")

    if not results.patient_results:
        st.warning("No individual patient data available")
        return

    sample_size = min(100, len(results.patient_results))
    sample_data = results.patient_results[:sample_size]

    trajectory_df = pd.DataFrame(sample_data)

    if not trajectory_df.empty:
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Outcome Distribution (Sample)**")
            if 'cumulative_costs' in trajectory_df.columns:
                st.bar_chart(trajectory_df['cumulative_costs'].head(50))
                st.caption("Cumulative Costs by Patient")

        with col2:
            if 'cumulative_qalys' in trajectory_df.columns:
                st.bar_chart(trajectory_df['cumulative_qalys'].head(50))
                st.caption("Cumulative QALYs by Patient")

        st.markdown("**Final State Distribution**")
        col3, col4 = st.columns(2)

        with col3:
            if 'cardiac_state' in trajectory_df.columns:
                cardiac_counts = trajectory_df['cardiac_state'].value_counts()
                st.dataframe(cardiac_counts.reset_index().rename(columns={'index': 'Cardiac State', 'cardiac_state': 'Count'}), hide_index=True)

        with col4:
            if 'renal_state' in trajectory_df.columns:
                renal_counts = trajectory_df['renal_state'].value_counts()
                st.dataframe(renal_counts.reset_index().rename(columns={'index': 'Renal State', 'renal_state': 'Count'}), hide_index=True)


def display_simulation_calculations(results: SimulationResults, currency: str):
    """Display detailed simulation calculations for sample patients."""
    st.markdown("### Microsimulation Calculations")
    st.markdown("""
    This section shows the detailed calculations applied during the microsimulation
    to randomize patient trajectories. View how transition probabilities are calculated
    and how events are sampled for individual patients.
    """)

    # Check if simulation log exists
    if not hasattr(results, 'simulation_log') or not results.simulation_log:
        st.warning("No detailed simulation log available. Run a new simulation to see calculations.")
        return

    sim_log = results.simulation_log

    # Patient selector
    patient_ids = list(sim_log.keys())
    selected_patient = st.selectbox(
        "Select Patient to View",
        patient_ids,
        format_func=lambda x: f"Patient {x} (Age {sim_log[x]['initial_age']:.0f}, SBP {sim_log[x]['initial_sbp']:.0f})"
    )

    patient_log = sim_log[selected_patient]

    # Patient summary
    st.markdown(f"""
    **Patient {selected_patient} Summary:**
    - Initial Age: {patient_log['initial_age']:.1f} years
    - Initial SBP: {patient_log['initial_sbp']:.0f} mmHg
    - Initial eGFR: {patient_log['initial_egfr']:.1f} mL/min/1.73m²
    - Treatment: {patient_log['treatment'].upper()}
    - Diabetes: {'Yes' if patient_log['has_diabetes'] else 'No'}
    - Heart Failure: {'Yes' if patient_log['has_hf'] else 'No'}
    """)

    if not patient_log['cycles']:
        st.info("No cycle data logged for this patient.")
        return

    # Create tabs for different views
    calc_tab1, calc_tab2, calc_tab3, calc_tab4 = st.tabs([
        "Transition Probabilities", "Trajectory Charts", "Event Log", "Calculation Details"
    ])

    # Convert cycles to dataframe
    cycles_df = pd.DataFrame(patient_log['cycles'])

    with calc_tab1:
        st.markdown("#### Monthly Transition Probabilities Over Time")
        st.markdown("""
        These probabilities are calculated each month using the PREVENT risk equations,
        modified by patient characteristics, prior events, and treatment effects.
        """)

        # Extract probability columns
        if 'probs' in cycles_df.columns:
            probs_df = pd.json_normalize(cycles_df['probs'])
            probs_df['year'] = cycles_df['year']

            # Display probability trends
            prob_cols = ['p_mi', 'p_ischemic_stroke', 'p_hf', 'p_cv_death', 'p_non_cv_death']
            prob_labels = ['MI', 'Ischemic Stroke', 'Heart Failure', 'CV Death', 'Non-CV Death']

            chart_data = probs_df[['year'] + prob_cols].copy()
            chart_data.columns = ['Year'] + prob_labels
            chart_data = chart_data.set_index('Year')

            st.line_chart(chart_data, use_container_width=True)
            st.caption("Monthly transition probabilities (higher = more likely)")

            # Show probability formula explanation
            with st.expander("How are these probabilities calculated?"):
                st.markdown("""
                **Transition Probability Calculation:**

                1. **Base Risk (PREVENT Equations)**
                   - Uses age, sex, SBP, eGFR, diabetes status, smoking, cholesterol, BMI
                   - Calculates 10-year risk, converted to monthly probability

                2. **Risk Multipliers Applied:**
                   - Prior MI: 2.5x MI risk
                   - Prior Stroke: 3.0x stroke risk
                   - Prior TIA: 2.0x stroke risk
                   - SGLT2i: 0.7x HF risk (30% reduction)

                3. **Monthly Probability Formula:**
                   ```
                   P_monthly = 1 - (1 - P_annual)^(1/12)
                   ```

                4. **Event Sampling:**
                   - Random number U ~ Uniform(0,1) drawn
                   - If U < P_event, event occurs
                   - Events are mutually exclusive (first match wins)
                """)

    with calc_tab2:
        st.markdown("#### Patient Trajectory Over Time")

        col1, col2 = st.columns(2)

        with col1:
            # SBP trajectory
            st.markdown("**Blood Pressure (SBP)**")
            bp_data = cycles_df[['year', 'sbp']].copy()
            bp_data.columns = ['Year', 'SBP (mmHg)']
            bp_data = bp_data.set_index('Year')
            st.line_chart(bp_data, use_container_width=True)

            st.markdown("""
            *SBP Update Equation:*
            ```
            SBP(t+1) = SBP(t) + 0.05 + ε - treatment_effect
            ```
            Where ε ~ N(0, 2) is monthly noise
            """)

        with col2:
            # eGFR trajectory
            st.markdown("**Renal Function (eGFR)**")
            egfr_data = cycles_df[['year', 'egfr']].copy()
            egfr_data.columns = ['Year', 'eGFR (mL/min)']
            egfr_data = egfr_data.set_index('Year')
            st.line_chart(egfr_data, use_container_width=True)

            st.markdown("""
            *eGFR Decline Equation:*
            ```
            eGFR(t+1) = eGFR(t) - base_decline - sbp_effect
            ```
            Where base_decline = 1-1.5 mL/min/year
            """)

        # Cumulative outcomes
        st.markdown("**Cumulative Outcomes**")
        col3, col4 = st.columns(2)

        with col3:
            costs_data = cycles_df[['year', 'cumulative_costs']].copy()
            costs_data.columns = ['Year', f'Costs ({currency})']
            costs_data = costs_data.set_index('Year')
            st.line_chart(costs_data, use_container_width=True)

        with col4:
            qalys_data = cycles_df[['year', 'cumulative_qalys']].copy()
            qalys_data.columns = ['Year', 'QALYs']
            qalys_data = qalys_data.set_index('Year')
            st.line_chart(qalys_data, use_container_width=True)

    with calc_tab3:
        st.markdown("#### Event Log")
        st.markdown("Events that occurred during the simulation for this patient:")

        # Filter for cycles with events
        events_df = cycles_df[cycles_df['event'].notna()].copy()

        if not events_df.empty:
            events_display = events_df[['year', 'event', 'sbp', 'egfr', 'cardiac_state']].copy()
            events_display.columns = ['Year', 'Event', 'SBP', 'eGFR', 'Resulting State']
            st.dataframe(events_display, use_container_width=True, hide_index=True)
        else:
            st.success("No adverse events occurred for this patient during the simulation.")

        # State transitions
        st.markdown("#### State Progression")
        states_df = cycles_df[['year', 'cardiac_state', 'renal_state', 'neuro_state', 'is_adherent']].copy()
        states_df.columns = ['Year', 'Cardiac', 'Renal', 'Cognitive', 'Adherent']
        st.dataframe(states_df, use_container_width=True, hide_index=True)

    with calc_tab4:
        st.markdown("#### Detailed Calculation at Each Time Point")
        st.markdown("Select a specific time point to see all calculations:")

        # Time point selector
        time_points = cycles_df['year'].tolist()
        selected_time = st.select_slider("Select Year", options=time_points, value=time_points[0] if time_points else 0)

        # Get data for selected time
        time_data = cycles_df[cycles_df['year'] == selected_time].iloc[0] if len(cycles_df[cycles_df['year'] == selected_time]) > 0 else None

        if time_data is not None:
            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**Patient State:**")
                st.markdown(f"""
                - Age: {time_data['age']:.1f} years
                - SBP (Office): {time_data['sbp']:.0f} mmHg
                - True SBP: {time_data['true_sbp']:.0f} mmHg
                - eGFR: {time_data['egfr']:.1f} mL/min
                - Adherent: {'Yes' if time_data['is_adherent'] else 'No'}
                - Treatment Effect: {time_data['treatment_effect']:.1f} mmHg/month
                """)

                st.markdown("**Current States:**")
                st.markdown(f"""
                - Cardiac: `{time_data['cardiac_state']}`
                - Renal: `{time_data['renal_state']}`
                - Cognitive: `{time_data['neuro_state']}`
                """)

            with col2:
                st.markdown("**Transition Probabilities (Monthly):**")
                probs = time_data['probs']
                prob_table = pd.DataFrame([
                    {'Event': 'MI', 'Probability': f"{probs['p_mi']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_mi'])**12)*100:.2f}%"},
                    {'Event': 'Ischemic Stroke', 'Probability': f"{probs['p_ischemic_stroke']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_ischemic_stroke'])**12)*100:.2f}%"},
                    {'Event': 'Hemorrhagic Stroke', 'Probability': f"{probs['p_hemorrhagic_stroke']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_hemorrhagic_stroke'])**12)*100:.2f}%"},
                    {'Event': 'TIA', 'Probability': f"{probs['p_tia']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_tia'])**12)*100:.2f}%"},
                    {'Event': 'Heart Failure', 'Probability': f"{probs['p_hf']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_hf'])**12)*100:.2f}%"},
                    {'Event': 'CV Death', 'Probability': f"{probs['p_cv_death']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_cv_death'])**12)*100:.2f}%"},
                    {'Event': 'Non-CV Death', 'Probability': f"{probs['p_non_cv_death']*100:.4f}%", 'Annual Equiv': f"{(1-(1-probs['p_non_cv_death'])**12)*100:.2f}%"},
                ])
                st.dataframe(prob_table, use_container_width=True, hide_index=True)

            st.markdown("**Outcome This Cycle:**")
            if time_data['event']:
                st.error(f"Event occurred: **{time_data['event']}**")
            else:
                st.success("No event - patient continues in current state")

            if time_data['adherence_changed']:
                st.warning("Adherence status changed this cycle")
            if time_data['neuro_changed']:
                st.warning("Cognitive state changed this cycle")
            if time_data['hyperkalemia_stop']:
                st.error("Treatment stopped due to hyperkalemia (K+ > 5.5)")


# ============== MAIN APPLICATION ==============

def main():
    """Main application entry point."""
    # Header
    st.markdown('<p class="main-header">Cost-Effectiveness Microsimulation</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">IXA-001 vs Spironolactone in Resistant Hypertension</p>', unsafe_allow_html=True)

    # ============== SIDEBAR ==============
    st.sidebar.markdown("## Simulation Parameters")

    n_patients = st.sidebar.slider("Cohort Size (per arm)", min_value=100, max_value=5000, value=1000, step=100)
    time_horizon = st.sidebar.slider("Time Horizon (years)", min_value=5, max_value=50, value=40, step=5)
    perspective = st.sidebar.selectbox("Cost Perspective", options=["US", "UK"], index=0)
    currency = "$" if perspective == "US" else "£"
    discount_rate = st.sidebar.slider("Discount Rate (%)", min_value=0.0, max_value=10.0, value=3.0, step=0.5) / 100.0
    seed = st.sidebar.number_input("Random Seed", min_value=1, max_value=99999, value=42)

    st.sidebar.markdown("---")
    st.sidebar.markdown("## Population Configuration")

    # Demographics
    with st.sidebar.expander("Demographics", expanded=False):
        age_mean = st.slider("Mean Age (years)", 40, 80, 62)
        age_sd = st.slider("Age SD", 5, 20, 10)
        age_min = st.slider("Age Min", 30, 60, 40)
        age_max = st.slider("Age Max", 70, 95, 85)
        prop_male = st.slider("% Male", 0, 100, 55) / 100.0
        bmi_mean = st.slider("Mean BMI", 20.0, 45.0, 30.5, step=0.5)
        bmi_sd = st.slider("BMI SD", 2.0, 10.0, 5.5, step=0.5)

    # Blood Pressure
    with st.sidebar.expander("Blood Pressure", expanded=False):
        sbp_mean = st.slider("Mean SBP (mmHg)", 140, 180, 155)
        sbp_sd = st.slider("SBP SD", 5, 25, 15)
        sbp_min = st.slider("SBP Min", 120, 150, 140)
        sbp_max = st.slider("SBP Max", 180, 220, 200)
        dbp_mean = st.slider("Mean DBP (mmHg)", 70, 110, 92)
        dbp_sd = st.slider("DBP SD", 5, 20, 10)

    # Renal Function
    with st.sidebar.expander("Renal Function", expanded=False):
        egfr_mean = st.slider("Mean eGFR", 30, 90, 68)
        egfr_sd = st.slider("eGFR SD", 10, 30, 20)
        egfr_min = st.slider("eGFR Min", 10, 30, 15)
        egfr_max = st.slider("eGFR Max", 90, 130, 120)
        uacr_mean = st.slider("Mean UACR (mg/g)", 10, 300, 50)
        uacr_sd = st.slider("UACR SD", 20, 150, 80)

    # Lipids
    with st.sidebar.expander("Lipids", expanded=False):
        total_chol_mean = st.slider("Mean Total Cholesterol", 150, 280, 200)
        total_chol_sd = st.slider("Total Cholesterol SD", 20, 60, 40)
        hdl_chol_mean = st.slider("Mean HDL Cholesterol", 30, 80, 48)
        hdl_chol_sd = st.slider("HDL Cholesterol SD", 5, 20, 12)

    # Standard Comorbidities
    with st.sidebar.expander("Standard Comorbidities (%)", expanded=False):
        diabetes_prev = st.slider("Diabetes", 0, 60, 35) / 100.0
        smoker_prev = st.slider("Current Smoker", 0, 40, 15) / 100.0
        dyslipidemia_prev = st.slider("Dyslipidemia", 0, 80, 60) / 100.0
        prior_mi_prev = st.slider("Prior MI", 0, 30, 10) / 100.0
        prior_stroke_prev = st.slider("Prior Stroke", 0, 20, 5) / 100.0
        heart_failure_prev = st.slider("Heart Failure", 0, 25, 8) / 100.0

    # Additional Comorbidities (NEW)
    with st.sidebar.expander("Additional Comorbidities (%)", expanded=False):
        st.caption("These conditions affect model dynamics")
        copd_prev = st.slider("COPD", 0, 40, 17) / 100.0
        depression_prev = st.slider("Depression", 0, 50, 27) / 100.0
        anxiety_prev = st.slider("Anxiety", 0, 40, 17) / 100.0
        substance_use_prev = st.slider("Substance Use Disorder", 0, 25, 10) / 100.0
        smi_prev = st.slider("Serious Mental Illness", 0, 15, 4) / 100.0
        afib_prev = st.slider("Atrial Fibrillation", 0, 25, 10) / 100.0
        pad_prev = st.slider("Peripheral Artery Disease", 0, 30, 15) / 100.0

    # Treatment & Adherence
    with st.sidebar.expander("Treatment & Adherence", expanded=False):
        adherence_prob = st.slider("Baseline Adherence (%)", 50, 100, 75) / 100.0
        mean_antihypertensives = st.slider("Mean Antihypertensives", 2, 6, 4)
        sglt2_uptake = st.slider("SGLT2i Uptake (%)", 0, 80, 40) / 100.0

    st.sidebar.markdown("---")
    st.sidebar.markdown("## Advanced Parameters")

    # Treatment Effects
    with st.sidebar.expander("Treatment Effects", expanded=False):
        st.markdown("**IXA-001**")
        ixa_sbp_reduction = st.slider("IXA-001 SBP Reduction (mmHg)", 10.0, 30.0, 20.0, step=1.0)
        ixa_sbp_sd = st.slider("IXA-001 SBP SD", 4.0, 12.0, 8.0, step=1.0)
        ixa_discontinuation = st.slider("IXA-001 Discontinuation (%/yr)", 5, 25, 12) / 100.0

        st.markdown("**Spironolactone**")
        spiro_sbp_reduction = st.slider("Spiro SBP Reduction (mmHg)", 5.0, 15.0, 9.0, step=1.0)
        spiro_sbp_sd = st.slider("Spiro SBP SD", 3.0, 10.0, 6.0, step=1.0)
        spiro_discontinuation = st.slider("Spiro Discontinuation (%/yr)", 10, 30, 15) / 100.0

        st.markdown("**Adherence Effect**")
        adherence_multiplier = st.slider("Non-Adherent Effect Multiplier", 0.1, 0.5, 0.3, step=0.05)

    # Clinical Parameters
    with st.sidebar.expander("Clinical Model Parameters", expanded=False):
        st.markdown("**Case Fatality Rates (30-day)**")
        cfr_mi = st.slider("MI CFR (%)", 2, 15, 8) / 100.0
        cfr_ischemic_stroke = st.slider("Ischemic Stroke CFR (%)", 5, 20, 10) / 100.0
        cfr_hemorrhagic_stroke = st.slider("Hemorrhagic Stroke CFR (%)", 15, 40, 25) / 100.0
        cfr_hf = st.slider("HF CFR (%)", 2, 12, 5) / 100.0

        st.markdown("**Stroke Distribution**")
        stroke_ischemic_frac = st.slider("Ischemic Stroke %", 70, 95, 85) / 100.0

        st.markdown("**Prior Event Risk Multipliers**")
        prior_mi_mult = st.slider("Prior MI Multiplier", 1.5, 4.0, 2.5, step=0.5)
        prior_stroke_mult = st.slider("Prior Stroke Multiplier", 2.0, 5.0, 3.0, step=0.5)

        st.markdown("**Cognitive Decline (Annual Rates)**")
        normal_to_mci = st.slider("Normal to MCI (%)", 1, 5, 2) / 100.0
        mci_to_dementia = st.slider("MCI to Dementia (%)", 5, 20, 10) / 100.0

        st.markdown("**Safety Monitoring**")
        hyperkalemia_threshold = st.slider("Hyperkalemia Threshold (K+)", 5.0, 6.5, 5.5, step=0.1)

    # Cost Parameters
    with st.sidebar.expander("Drug Costs (Monthly)", expanded=False):
        ixa_monthly_cost = st.number_input("IXA-001", value=500.0 if perspective == "US" else 400.0, step=50.0)
        spiro_monthly_cost = st.number_input("Spironolactone", value=15.0 if perspective == "US" else 8.0, step=5.0)
        sglt2_monthly_cost = st.number_input("SGLT2 Inhibitor", value=450.0 if perspective == "US" else 35.0, step=25.0)
        background_monthly_cost = st.number_input("Background Therapy", value=75.0 if perspective == "US" else 40.0, step=10.0)
        lab_cost_k = st.number_input("K+ Lab Test", value=15.0 if perspective == "US" else 3.0, step=5.0)

    with st.sidebar.expander("Acute Event Costs", expanded=False):
        mi_acute_cost = st.number_input("MI (Acute)", value=25000.0 if perspective == "US" else 8000.0, step=1000.0)
        ischemic_stroke_cost = st.number_input("Ischemic Stroke", value=15200.0 if perspective == "US" else 6000.0, step=500.0)
        hemorrhagic_stroke_cost = st.number_input("Hemorrhagic Stroke", value=22500.0 if perspective == "US" else 9000.0, step=500.0)
        tia_cost = st.number_input("TIA", value=2100.0 if perspective == "US" else 850.0, step=100.0)
        hf_admission_cost = st.number_input("HF Admission", value=18000.0 if perspective == "US" else 5500.0, step=1000.0)

    with st.sidebar.expander("Annual Management Costs", expanded=False):
        controlled_htn_annual = st.number_input("Controlled HTN", value=800.0 if perspective == "US" else 350.0, step=100.0)
        uncontrolled_htn_annual = st.number_input("Uncontrolled HTN", value=1200.0 if perspective == "US" else 550.0, step=100.0)
        post_mi_annual = st.number_input("Post-MI", value=5500.0 if perspective == "US" else 2200.0, step=500.0)
        post_stroke_annual = st.number_input("Post-Stroke", value=12000.0 if perspective == "US" else 5500.0, step=500.0)
        hf_annual = st.number_input("Heart Failure", value=15000.0 if perspective == "US" else 6000.0, step=1000.0)
        ckd_3a_annual = st.number_input("CKD Stage 3a", value=2500.0 if perspective == "US" else 1200.0, step=250.0)
        ckd_3b_annual = st.number_input("CKD Stage 3b", value=4500.0 if perspective == "US" else 2200.0, step=250.0)
        ckd_4_annual = st.number_input("CKD Stage 4", value=8000.0 if perspective == "US" else 3500.0, step=500.0)
        esrd_annual = st.number_input("ESRD", value=90000.0 if perspective == "US" else 35000.0, step=5000.0)

    with st.sidebar.expander("Indirect Costs (Productivity)", expanded=False):
        daily_wage = st.number_input("Daily Wage", value=240.0 if perspective == "US" else 160.0, step=20.0)
        absenteeism_mi = st.number_input("Absenteeism MI (days)", value=7 if perspective == "US" else 14, step=1)
        absenteeism_stroke = st.number_input("Absenteeism Stroke (days)", value=30 if perspective == "US" else 60, step=5)
        absenteeism_hf = st.number_input("Absenteeism HF (days)", value=5 if perspective == "US" else 10, step=1)
        disability_stroke = st.slider("Disability Multiplier Stroke (%)", 10, 50, 20 if perspective == "US" else 30) / 100.0
        disability_hf = st.slider("Disability Multiplier HF (%)", 5, 35, 15 if perspective == "US" else 20) / 100.0

    # Utility Parameters
    with st.sidebar.expander("Utility/QALY Parameters", expanded=False):
        st.markdown("**Baseline Utilities by Age**")
        util_40 = st.slider("Age 40 Utility", 0.70, 1.0, 0.90, step=0.01)
        util_60 = st.slider("Age 60 Utility", 0.65, 0.95, 0.84, step=0.01)
        util_80 = st.slider("Age 80 Utility", 0.55, 0.90, 0.75, step=0.01)

        st.markdown("**Disutilities (Decrements)**")
        disutil_uncontrolled = st.slider("Uncontrolled HTN", 0.0, 0.10, 0.04, step=0.01)
        disutil_post_mi = st.slider("Post-MI", 0.05, 0.25, 0.12, step=0.01)
        disutil_post_stroke = st.slider("Post-Stroke", 0.10, 0.35, 0.18, step=0.01)
        disutil_esrd = st.slider("ESRD", 0.20, 0.50, 0.35, step=0.01)
        disutil_diabetes = st.slider("Diabetes", 0.0, 0.10, 0.04, step=0.01)

    # Build population params
    pop_params = PopulationParams(
        n_patients=n_patients, seed=seed,
        age_mean=age_mean, age_sd=age_sd, age_min=age_min, age_max=age_max,
        prop_male=prop_male,
        sbp_mean=sbp_mean, sbp_sd=sbp_sd, sbp_min=sbp_min, sbp_max=sbp_max,
        dbp_mean=dbp_mean, dbp_sd=dbp_sd,
        egfr_mean=egfr_mean, egfr_sd=egfr_sd, egfr_min=egfr_min, egfr_max=egfr_max,
        uacr_mean=uacr_mean, uacr_sd=uacr_sd,
        total_chol_mean=total_chol_mean, total_chol_sd=total_chol_sd,
        hdl_chol_mean=hdl_chol_mean, hdl_chol_sd=hdl_chol_sd,
        bmi_mean=bmi_mean, bmi_sd=bmi_sd,
        diabetes_prev=diabetes_prev, smoker_prev=smoker_prev,
        dyslipidemia_prev=dyslipidemia_prev,
        prior_mi_prev=prior_mi_prev, prior_stroke_prev=prior_stroke_prev,
        heart_failure_prev=heart_failure_prev,
        adherence_prob=adherence_prob,
        mean_antihypertensives=mean_antihypertensives,
    )

    # Build treatment params
    treatment_params = TreatmentParams(
        ixa_sbp_reduction=ixa_sbp_reduction,
        ixa_sbp_reduction_sd=ixa_sbp_sd,
        ixa_discontinuation_rate=ixa_discontinuation,
        spiro_sbp_reduction=spiro_sbp_reduction,
        spiro_sbp_reduction_sd=spiro_sbp_sd,
        spiro_discontinuation_rate=spiro_discontinuation,
        adherence_effect_multiplier=adherence_multiplier,
    )

    # Build clinical params
    clinical_params = ClinicalParams(
        cfr_mi=cfr_mi,
        cfr_ischemic_stroke=cfr_ischemic_stroke,
        cfr_hemorrhagic_stroke=cfr_hemorrhagic_stroke,
        cfr_hf=cfr_hf,
        stroke_ischemic_fraction=stroke_ischemic_frac,
        prior_mi_multiplier=prior_mi_mult,
        prior_stroke_multiplier=prior_stroke_mult,
        normal_to_mci_rate=normal_to_mci,
        mci_to_dementia_rate=mci_to_dementia,
    )

    # Build custom costs
    custom_costs = CustomCostInputs(
        ixa_001_monthly=ixa_monthly_cost,
        spironolactone_monthly=spiro_monthly_cost,
        sglt2_inhibitor_monthly=sglt2_monthly_cost,
        background_therapy_monthly=background_monthly_cost,
        lab_test_cost_k=lab_cost_k,
        mi_acute=mi_acute_cost,
        ischemic_stroke_acute=ischemic_stroke_cost,
        hemorrhagic_stroke_acute=hemorrhagic_stroke_cost,
        tia_acute=tia_cost,
        hf_admission=hf_admission_cost,
        controlled_htn_annual=controlled_htn_annual,
        uncontrolled_htn_annual=uncontrolled_htn_annual,
        post_mi_annual=post_mi_annual,
        post_stroke_annual=post_stroke_annual,
        heart_failure_annual=hf_annual,
        ckd_stage_3a_annual=ckd_3a_annual,
        ckd_stage_3b_annual=ckd_3b_annual,
        ckd_stage_4_annual=ckd_4_annual,
        esrd_annual=esrd_annual,
        daily_wage=daily_wage,
        absenteeism_mi_days=absenteeism_mi,
        absenteeism_stroke_days=absenteeism_stroke,
        absenteeism_hf_days=absenteeism_hf,
        disability_multiplier_stroke=disability_stroke,
        disability_multiplier_hf=disability_hf,
    )

    # Build utility params
    utility_params = UtilityParams(
        baseline_utility_40=util_40,
        baseline_utility_60=util_60,
        baseline_utility_80=util_80,
        disutility_uncontrolled_htn=disutil_uncontrolled,
        disutility_post_mi=disutil_post_mi,
        disutility_post_stroke=disutil_post_stroke,
        disutility_esrd=disutil_esrd,
        disutility_diabetes=disutil_diabetes,
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### About")
    st.sidebar.markdown("""
    This microsimulation model evaluates the cost-effectiveness of
    **IXA-001** (aldosterone synthase inhibitor) compared to
    **Spironolactone** in adults with resistant hypertension.

    **All model parameters are now configurable** including:
    - Population demographics & comorbidities
    - Treatment effects & discontinuation
    - Clinical event rates & fatality
    - All cost parameters
    - Utility/QALY values
    """)

    # Run simulation button
    if st.sidebar.button("Run Simulation", type="primary", use_container_width=True):
        with st.status(f"Running microsimulation ({n_patients:,} patients per arm, {time_horizon} years)...", expanded=True) as status:
            cea_results, patients_ixa, patients_spi, profiles = run_simulation_with_progress(
                n_patients, time_horizon, perspective, seed, discount_rate, pop_params, status,
                custom_costs, treatment_params, clinical_params
            )
            st.session_state.cea_results = cea_results
            st.session_state.currency = currency
            st.session_state.pop_params = pop_params
            st.session_state.patients_ixa = patients_ixa
            st.session_state.profiles = profiles
            st.session_state.subgroup_data = analyze_subgroups(patients_ixa, cea_results.intervention, profiles)
            st.session_state.custom_costs = custom_costs
            st.session_state.treatment_params = treatment_params
            st.session_state.clinical_params = clinical_params
            # Store simulation logs for calculation visualization
            st.session_state.sim_log_ixa = getattr(cea_results.intervention, 'simulation_log', {})
            st.session_state.sim_log_spi = getattr(cea_results.comparator, 'simulation_log', {})

    # ============== MAIN CONTENT ==============
    if "cea_results" in st.session_state:
        cea = st.session_state.cea_results
        currency = st.session_state.currency
        pp = st.session_state.pop_params
        profiles = st.session_state.profiles
        subgroup_data = st.session_state.subgroup_data
        custom_costs = st.session_state.get('custom_costs')
        treatment_params = st.session_state.get('treatment_params')
        clinical_params = st.session_state.get('clinical_params')

        # Key metrics
        display_key_metrics(cea, currency)

        st.divider()

        # Tabs for different views
        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            "Outcomes", "Costs", "Charts", "Subgroups", "Risk Stratification", "Trajectories", "Calculations", "Export"
        ])

        with tab1:
            display_outcomes_table(cea, currency)
            st.divider()
            display_medication_adherence(cea)

        with tab2:
            display_detailed_costs(cea, currency)
            st.divider()
            display_wtp_analysis(cea, currency)

        with tab3:
            display_event_charts(cea)
            st.divider()
            display_ce_plane(cea, currency)

        with tab4:
            display_subgroup_analysis(subgroup_data, currency)

        with tab5:
            display_risk_stratification(profiles)

        with tab6:
            display_patient_trajectories(st.session_state.patients_ixa, cea.intervention)

        with tab7:
            # Simulation Calculations tab
            st.markdown("### Simulation Calculations")
            calc_arm = st.radio("Select Treatment Arm", ["IXA-001", "Spironolactone"], horizontal=True)

            if calc_arm == "IXA-001":
                display_simulation_calculations(cea.intervention, currency)
            else:
                display_simulation_calculations(cea.comparator, currency)

        with tab8:
            st.markdown("### Export Results")
            st.markdown("Download comprehensive Excel report with all analysis results and parameters used.")

            excel_buffer = generate_excel_report(cea, pp, subgroup_data, currency, custom_costs, treatment_params, clinical_params)
            st.download_button(
                label="Download Excel Report",
                data=excel_buffer,
                file_name="CEA_Microsimulation_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

            st.markdown("""
            **Report Contents:**
            - Executive Summary with key results
            - Clinical Events comparison with charts
            - Cost Analysis (direct & indirect)
            - Subgroup Analysis by risk categories
            - Willingness-to-Pay Analysis
            - All Simulation Parameters used
            """)

        # Summary box
        st.divider()
        st.markdown("### Summary")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown(f"""
            **Simulation Parameters:**
            - Cohort size: {n_patients:,} patients per arm
            - Time horizon: {time_horizon} years
            - Perspective: {perspective}
            - Discount rate: {discount_rate*100:.1f}% per annum
            """)

        with col2:
            st.markdown(f"""
            **Population Characteristics:**
            - Mean age: {pp.age_mean:.0f} years (SD {pp.age_sd:.0f})
            - Male: {pp.prop_male*100:.0f}%
            - Mean SBP: {pp.sbp_mean:.0f} mmHg
            - Mean eGFR: {pp.egfr_mean:.0f} mL/min
            - Diabetes: {pp.diabetes_prev*100:.0f}%
            """)

        with col3:
            events_avoided = cea.comparator.stroke_events - cea.intervention.stroke_events
            mi_avoided = cea.comparator.mi_events - cea.intervention.mi_events

            st.markdown(f"""
            **Key Findings:**
            - Strokes avoided: {events_avoided:,}
            - MIs avoided: {mi_avoided:,}
            - Additional QALYs: {cea.incremental_qalys:.2f}
            - Additional cost: {currency}{cea.incremental_costs:,.0f}
            """)

    else:
        st.info("Configure parameters and click **Run Simulation** to start the analysis.")


if __name__ == "__main__":
    main()
